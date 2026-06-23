"""
workbook.py - Classeur Excel enrichi : analyse DAF, reporting et rapport de réconciliation.

Présentation Excel pure : ce module assemble et met en forme le classeur, sans aucune
logique métier (toute la logique vit dans utils.py).

Fonctions publiques :
    compute_dashboard_kpis()              → indicateurs clés pré-calculés
    build_excel_bytes_with_dashboard()    → export mode Streamlit (en mémoire)
    export_excel_with_dashboard()         → export mode Batch (sur disque)

Structure du classeur produit (9 feuilles, dans l'ordre) :
    1. 🧭 Synthèse DAF          → narratif factuel + indicateurs clés
    2. 💹 Marge & rentabilité   → érosion mensuelle, mix catégorie, fuite par remise
    3. 🎯 Réalisé vs Budget     → écarts au plan par mois et par région
    4. 👥 Clients & segments    → concentration (Pareto), top clients, poids des segments
    5. 📊 Dashboard             → KPIs + graphiques (matplotlib en PNG)
    6. 📅 Par mois              → agrégation mensuelle formatée
    7. 👤 Par commercial        → CA, marge et taux par commercial
    8. 🔍 Annexe (données brutes) → transactions livrées nettoyées (autofiltre)
    9. 🔐 Réconciliation        → miroir Excel du rapport d'intégrité (OK/ALERTE)

Palette : bleu marine (#1F3864), gris, blanc — sobre, adapté direction financière PME.
Graphiques : matplotlib (PNG insérés via openpyxl) — titres et axes hors de la zone de
    tracé, graduations complètes avec formatage monétaire.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
plt.switch_backend("agg")   # backend non-interactif — requis en environnement serveur (Streamlit)
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from config import SETTINGS
from utils import ReconciliationReport

# ===========================================================================
# PALETTE DE COULEURS — Charte sobre (bleu marine / gris / blanc)
# ===========================================================================
# Hexadécimal sans '#', format openpyxl.

_NAVY       = "1F3864"   # bleu marine foncé  — titres, valeurs KPI
_NAVY_MID   = "2E75B6"   # bleu moyen         — accents, courbes secondaires
_GREY_DARK  = "404040"   # gris foncé         — texte courant
_GREY_MID   = "7F7F7F"   # gris moyen         — libellés secondaires
_GREY_LITE  = "F2F2F2"   # gris très clair    — fond cartes KPI, alternance lignes
_BORDER_COL = "D9D9D9"   # gris clair         — bordures
_WHITE      = "FFFFFF"   # blanc              — fond principal
_GREEN      = "375623"   # vert foncé sobre   — variation positive
_RED        = "C00000"   # rouge foncé sobre  — variation négative / alerte
_ALERT_BG   = "FFF2CC"   # jaune très pâle    — fond bloc alerte réconciliation

# Dimensions des graphiques (cm)
_CHART_W = 13.5
_CHART_H = 10.5

# Largeurs des colonnes de la feuille Dashboard
_COL_WIDTHS_DB: dict[int, float] = {
    1: 20,   # A — KPI col 1
    2: 3,    # B — espaceur
    3: 20,   # C — KPI col 2
    4: 3,    # D — espaceur
    5: 20,   # E — KPI col 3
    6: 3,    # F — espaceur droit
}

# ===========================================================================
# STYLES RÉUTILISABLES
# ===========================================================================

_THIN_SIDE   = Side(border_style="thin", color=_BORDER_COL)
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)


def _font(
    bold: bool = False,
    size: int = 10,
    color: str = _GREY_DARK,
    italic: bool = False,
) -> Font:
    """Construit un objet Font Calibri avec les paramètres donnés."""
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def _fill(color: str) -> PatternFill:
    """Construit un PatternFill uni."""
    return PatternFill(fill_type="solid", fgColor=color)


# ===========================================================================
# HELPER D'ÉCRITURE DE CELLULE
# ===========================================================================

def _w(
    ws: Any,
    row: int,
    col: int,
    value: Any = None,
    *,
    bold: bool = False,
    size: int = 10,
    color: str = _GREY_DARK,
    italic: bool = False,
    bg: Optional[str] = None,
    align: str = "left",
    wrap: bool = False,
    num_fmt: str = "General",
) -> Any:
    """Écrit une valeur dans une cellule et applique le style demandé.

    Fonction volontairement courte (nommée ``_w``) car appelée très souvent
    dans les builders de feuilles. Ne gère pas les bordures, qui sont appliquées
    manuellement quand un contrôle précis par côté est nécessaire.

    Args:
        ws: Feuille openpyxl cible.
        row: Ligne (1-indexé).
        col: Colonne (1-indexé).
        value: Valeur à écrire (None = ne pas modifier la valeur).
        bold: Gras.
        size: Taille de police.
        color: Couleur hex de la police.
        italic: Italique.
        bg: Couleur hex de fond (None = pas de fond).
        align: Alignement horizontal ("left", "center", "right").
        wrap: Retour à la ligne automatique.
        num_fmt: Format numérique Excel (ex: ``'#,##0.00 "€"'``).

    Returns:
        La cellule modifiée.
    """
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font      = _font(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = _fill(bg)
    if num_fmt != "General":
        cell.number_format = num_fmt
    return cell


def _fmt_eur(x: float, decimals: int = 0) -> str:
    """Formate un montant en euros avec séparateur de milliers français.

    Args:
        x: Montant numérique.
        decimals: Nombre de décimales (0 ou 2).

    Returns:
        Chaîne formatée, ex: ``"1 234 567 €"`` ou ``"1 234,56 €"``.
    """
    if decimals == 0:
        return f"{x:,.0f} €".replace(",", " ")
    s = f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
    return f"{s} €"


def _fmt_signed_pct(x: Optional[float], decimals: int = 1) -> str:
    """Formate un pourcentage signé à la française : ``16.9`` → ``+16,9 %`` ; ``None`` → ``n/a``."""
    if x is None or pd.isna(x):
        return "n/a"
    return f"{x:+.{decimals}f} %".replace(".", ",")


def _fmt_signed_pts(x: Optional[float], decimals: int = 1) -> str:
    """Formate une variation en points signée : ``-2.12`` → ``-2,1 pts`` ; ``None`` → ``n/a``."""
    if x is None or pd.isna(x):
        return "n/a"
    return f"{x:+.{decimals}f} pts".replace(".", ",")


def _favorable_color(value: Optional[float], higher_is_better: bool = True) -> str:
    """Renvoie la couleur sémantique d'une variation : vert si favorable, rouge sinon.

    Un ``None`` (variation indéfinie, ex. YoY sur une seule année) reste neutre (navy).
    """
    if value is None or pd.isna(value):
        return _NAVY
    favorable = value >= 0 if higher_is_better else value <= 0
    return _GREEN if favorable else _RED


# ===========================================================================
# DATACLASS — Indicateurs clés du tableau de bord
# ===========================================================================

@dataclass
class DashboardKPIs:
    """Indicateurs clés pré-calculés pour le tableau de bord DAF.

    Attributes:
        total_revenue: CA total de la période (€).
        transaction_count: Nombre de transactions avec montant valide.
        average_basket: Panier moyen en € (CA / nb transactions valides).
        monthly_avg_revenue: CA moyen par mois.
        active_months: Nombre de mois distincts couverts.
        best_month: Mois YYYY-MM au CA le plus élevé.
        best_month_revenue: CA du meilleur mois.
        worst_month: Mois YYYY-MM au CA le plus faible.
        worst_month_revenue: CA du pire mois.
        last_mom_growth_pct: Variation M/M % entre les 2 derniers mois.
            None si données insuffisantes (< 2 mois).
        top_salesperson: Commercial avec le plus haut CA. None si absent.
        top_salesperson_revenue: CA du top commercial.
        active_salespeople: Nombre de commerciaux distincts.
    """

    total_revenue: float
    transaction_count: int
    average_basket: float
    monthly_avg_revenue: float
    active_months: int
    best_month: str
    best_month_revenue: float
    worst_month: str
    worst_month_revenue: float
    last_mom_growth_pct: Optional[float]
    top_salesperson: Optional[str]
    top_salesperson_revenue: Optional[float]
    active_salespeople: int


def compute_dashboard_kpis(
    df: pd.DataFrame,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
) -> DashboardKPIs:
    """Calcule les indicateurs clés à partir des données nettoyées.

    Tous les calculs sont basés sur les montants valides (non-NaN) uniquement,
    conformément à la règle comptable : un montant non convertible n'est pas du CA.

    Args:
        df: DataFrame nettoyé (colonnes : date, montant, mois, commercial optionnel).
        report_months: Agrégation mensuelle [mois, montant], non triée.
        report_salespeople: Agrégation par commercial [commercial, montant] ou vide.

    Returns:
        DashboardKPIs avec tous les indicateurs calculés.
    """
    valid_amounts     = df["montant"].dropna()
    total_revenue     = float(valid_amounts.sum())
    transaction_count = int(valid_amounts.count())
    average_basket    = total_revenue / transaction_count if transaction_count > 0 else 0.0

    months_sorted     = report_months.sort_values("mois").reset_index(drop=True)
    active_months     = int(months_sorted["mois"].nunique())
    monthly_avg       = total_revenue / active_months if active_months > 0 else 0.0

    idx_best  = int(months_sorted["montant"].idxmax())
    idx_worst = int(months_sorted["montant"].idxmin())
    best_month          = str(months_sorted.loc[idx_best,  "mois"])
    best_month_revenue  = float(months_sorted.loc[idx_best,  "montant"])
    worst_month         = str(months_sorted.loc[idx_worst, "mois"])
    worst_month_revenue = float(months_sorted.loc[idx_worst, "montant"])

    # Variation M/M : compare les 2 derniers mois connus.
    # Justification métier : on ne calcule pas sur la moyenne pour éviter
    # de noyer une rupture récente dans l'historique.
    last_mom_growth_pct: Optional[float] = None
    if len(months_sorted) >= 2:
        last_val = float(months_sorted["montant"].iloc[-1])
        prev_val = float(months_sorted["montant"].iloc[-2])
        if prev_val != 0:
            last_mom_growth_pct = (last_val - prev_val) / abs(prev_val) * 100.0

    top_salesperson:         Optional[str]   = None
    top_salesperson_revenue: Optional[float] = None
    active_salespeople = 0

    if not report_salespeople.empty:
        active_salespeople = int(report_salespeople["commercial"].nunique())
        idx_top = int(report_salespeople["montant"].idxmax())
        top_salesperson         = str(report_salespeople.loc[idx_top, "commercial"])
        top_salesperson_revenue = float(report_salespeople.loc[idx_top, "montant"])

    return DashboardKPIs(
        total_revenue=total_revenue,
        transaction_count=transaction_count,
        average_basket=average_basket,
        monthly_avg_revenue=monthly_avg,
        active_months=active_months,
        best_month=best_month,
        best_month_revenue=best_month_revenue,
        worst_month=worst_month,
        worst_month_revenue=worst_month_revenue,
        last_mom_growth_pct=last_mom_growth_pct,
        top_salesperson=top_salesperson,
        top_salesperson_revenue=top_salesperson_revenue,
        active_salespeople=active_salespeople,
    )



# ===========================================================================
# GRAPHIQUES DU DASHBOARD — Builders individuels
# ===========================================================================


# ===========================================================================
# HELPERS MATPLOTLIB — Formatage et style commun
# ===========================================================================

def _eur_axis_fmt(x: float, _: Any) -> str:
    """Formatte une valeur monétaire pour l'affichage sur un axe matplotlib.

    Utilise les suffixes k et M pour les grands montants afin d'éviter
    l'encombrement de l'axe avec des chiffres complets.

    Args:
        x: Valeur numérique.
        _: Position (ignorée, requise par l'API FuncFormatter de matplotlib).

    Returns:
        Chaîne formatée, ex: ``"1 250k €"``, ``"2.3M €"``, ``"850 €"``.
    """
    if abs(x) >= 1_000_000:
        return f"{x / 1_000_000:.1f}M €"
    if abs(x) >= 1_000:
        return f"{x / 1_000:.0f}k €"
    return f"{x:.0f} €"


def _pct_axis_fmt(x: float, _: Any) -> str:
    """Formatte une valeur en pourcentage pour l'affichage sur un axe matplotlib."""
    return f"{x:+.1f} %" if x != 0 else "0 %"


def _pct_axis_fmt_unsigned(x: float, _: Any) -> str:
    """Formatte un pourcentage SANS signe — pour les taux (niveaux), pas les variations."""
    return f"{x:.0f} %"


def _apply_chart_style(
    ax: Any,
    title: str,
    xlabel: str,
    ylabel: str,
    grid_axis: str = "y",
) -> None:
    """Applique la charte graphique sobre (navy/gris/blanc) à un axe matplotlib.

    Le titre est placé AU-DESSUS de la zone de tracé (comportement natif matplotlib),
    les labels d'axes sont À L'EXTÉRIEUR par défaut — ce qui résout le problème
    de chevauchement des graphiques openpyxl natifs.

    Args:
        ax: Axe matplotlib à styliser.
        title: Titre du graphique (au-dessus, aligné à gauche, navy gras).
        xlabel: Label de l'axe X (en dessous, gris).
        ylabel: Label de l'axe Y (à gauche, gris).
        grid_axis: Axe sur lequel afficher la grille (``"x"``, ``"y"`` ou ``"both"``).
    """
    ax.set_title(
        title,
        fontsize=10, fontweight="bold",
        color=f"#{_NAVY}", loc="left", pad=10,
    )
    ax.set_xlabel(xlabel, fontsize=8, color=f"#{_GREY_MID}", labelpad=6)
    ax.set_ylabel(ylabel, fontsize=8, color=f"#{_GREY_MID}", labelpad=6)
    ax.tick_params(
        axis="both", labelsize=7.5,
        labelcolor=f"#{_GREY_DARK}", length=3, width=0.5,
    )
    ax.grid(
        axis=grid_axis,
        color=f"#{_BORDER_COL}", linestyle="-", linewidth=0.5, alpha=0.8,
    )
    ax.set_axisbelow(True)
    ax.set_facecolor(f"#{_GREY_LITE}")
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("left", "bottom"):
        ax.spines[spine].set_color(f"#{_BORDER_COL}")
        ax.spines[spine].set_linewidth(0.5)


def _fig_to_xl_image(fig: Any) -> XLImage:
    """Convertit une figure matplotlib en objet Image openpyxl (PNG en mémoire).

    La figure est fermée après conversion pour libérer la mémoire (important en
    mode Streamlit où plusieurs rapports peuvent être générés en séquence).
    La taille d'affichage dans Excel est alignée sur ``_CHART_W`` / ``_CHART_H``.

    Args:
        fig: Figure matplotlib à convertir.

    Returns:
        Objet ``XLImage`` prêt à être inséré avec ``ws.add_image()``.
    """
    buf = io.BytesIO()
    fig.savefig(
        buf, format="png", dpi=150,
        bbox_inches="tight", facecolor=fig.get_facecolor(),
    )
    plt.close(fig)
    buf.seek(0)
    img        = XLImage(buf)
    img.width  = int(_CHART_W / 2.54 * 96)   # cm → pixels (96 dpi référence Excel)
    img.height = int(_CHART_H / 2.54 * 96)
    return img


def _chart_ca_mensuel(months: pd.DataFrame) -> XLImage:
    """Graphique 1 — CA mensuel (barres verticales, bleu marine).

    Le graphique le plus fondamental pour un DAF. Permet de lire d'un coup
    d'œil les mois forts et les creux d'activité, base de toute discussion
    budgétaire ou de prévision.

    Args:
        months: DataFrame mensuel trié par mois [mois, montant].

    Returns:
        Image PNG openpyxl prête à être insérée dans la feuille Excel.
    """
    labels = months["mois"].astype(str).tolist()
    values = months["montant"].tolist()
    x      = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.bar(x, values, color=f"#{_NAVY}", width=0.6, zorder=3)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))

    _apply_chart_style(ax, "Chiffre d'affaires mensuel", "Mois", "Montant (€)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_ca_cumule(months: pd.DataFrame) -> XLImage:
    """Graphique 2 — CA cumulé sur la période (courbe avec remplissage, bleu moyen).

    La courbe cumulée révèle la trajectoire de l'exercice : accélération,
    stagnation ou ralentissement. Indispensable pour les projections de
    fin d'année et la comparaison N vs N-1 (si plusieurs exercices).

    Args:
        months: DataFrame mensuel trié par mois [mois, montant].

    Returns:
        Image PNG openpyxl.
    """
    labels = months["mois"].astype(str).tolist()
    cumsum: list[float] = []
    s = 0.0
    for v in months["montant"].tolist():
        s += v
        cumsum.append(s)
    x = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.plot(list(x), cumsum, color=f"#{_NAVY_MID}", linewidth=2, zorder=3)
    ax.scatter(list(x), cumsum, color=f"#{_NAVY_MID}", s=25, zorder=4)
    ax.fill_between(list(x), cumsum, alpha=0.12, color=f"#{_NAVY_MID}")
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))

    _apply_chart_style(ax, "CA cumulé sur la période", "Mois", "Montant cumulé (€)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_variation_mom(months: pd.DataFrame) -> XLImage:
    """Graphique 3 — Variation mensuelle M/M en % (barres, vert/rouge selon signe).

    Signal d'alerte immédiat : détecte les ruptures de tendance mois par mois.
    Les barres sont colorées en vert (hausse) ou rouge (baisse) pour une lecture
    instantanée — avantage clé de matplotlib vs openpyxl natif qui ne supporte
    pas les couleurs conditionnelles par valeur.
    Le premier mois est exclu (pas de mois précédent disponible).

    Args:
        months: DataFrame mensuel trié par mois [mois, montant].

    Returns:
        Image PNG openpyxl.
    """
    raw_values = months["montant"].tolist()
    raw_labels = months["mois"].astype(str).tolist()

    # Calcul des variations — le premier mois n'a pas de précédent
    variations: list[float] = []
    v_labels:   list[str]   = []
    for i in range(1, len(raw_values)):
        if raw_values[i - 1] != 0:
            variations.append((raw_values[i] - raw_values[i - 1]) / abs(raw_values[i - 1]) * 100.0)
            v_labels.append(raw_labels[i])

    x      = range(len(v_labels))
    colors = [f"#{_GREEN}" if v >= 0 else f"#{_RED}" for v in variations]

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.bar(list(x), variations, color=colors, width=0.6, zorder=3)
    ax.axhline(y=0, color=f"#{_GREY_MID}", linewidth=0.8, zorder=2)
    ax.set_xticks(list(x))
    ax.set_xticklabels(v_labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_pct_axis_fmt))

    _apply_chart_style(ax, "Variation mensuelle M/M (%)", "Mois", "Variation (%)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_top_commerciaux(salespeople: pd.DataFrame) -> Optional[XLImage]:
    """Graphique 4 — Classement des commerciaux par CA (barres horizontales).

    Barres horizontales pour afficher les noms complets sans troncature.
    Les noms sont inversés pour que le meilleur apparaisse en haut.
    Limité aux 10 premiers pour la lisibilité.

    Args:
        salespeople: DataFrame [commercial, montant] trié par CA décroissant,
            ou DataFrame vide si pas de données commerciales.

    Returns:
        Image PNG openpyxl, ou None si ``salespeople`` est vide.
    """
    if salespeople.empty:
        return None

    top    = salespeople.head(10).reset_index(drop=True)
    # Inversion pour que le top 1 soit en haut du graphique
    labels = top["commercial"].astype(str).tolist()[::-1]
    values = top["montant"].tolist()[::-1]
    y      = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.barh(list(y), values, color=f"#{_NAVY_MID}", height=0.6, zorder=3)
    ax.set_yticks(list(y))
    ax.set_yticklabels(labels)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))

    _apply_chart_style(
        ax, "Classement des commerciaux par CA",
        "Montant (€)", "Commercial", grid_axis="x",
    )
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_panier_moyen(months: pd.DataFrame, df: pd.DataFrame) -> XLImage:
    """Graphique 5 — Évolution du panier moyen mensuel (courbe marqueurs diamant).

    Indicateur qualité stratégiquement clé : un CA stable avec un panier en hausse
    signifie moins de clients mais mieux valorisés → signal de montée en gamme.
    Inversement, un panier en baisse avec CA stable signale une dilution du portefeuille.

    Args:
        months: DataFrame mensuel trié par mois [mois, montant].
        df: DataFrame nettoyé complet, pour calculer le nb de transactions par mois.

    Returns:
        Image PNG openpyxl.
    """
    tx_by_month = (
        df[df["montant"].notna()]
        .groupby("mois").size()
        .rename("nb_tx")
        .reset_index()
    )
    data          = months.merge(tx_by_month, on="mois", how="left")
    data["nb_tx"] = data["nb_tx"].fillna(1).clip(lower=1)
    data["panier"] = data["montant"] / data["nb_tx"]

    labels = data["mois"].astype(str).tolist()
    values = data["panier"].tolist()
    x      = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.plot(list(x), values, color=f"#{_GREY_DARK}", linewidth=2, zorder=3)
    ax.scatter(list(x), values, color=f"#{_GREY_DARK}", s=30, marker="D", zorder=4)
    ax.fill_between(list(x), values, alpha=0.08, color=f"#{_GREY_DARK}")
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))

    _apply_chart_style(ax, "Panier moyen mensuel (€ / transaction)", "Mois", "Panier moyen (€)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


# ===========================================================================
# FEUILLE DASHBOARD — Mise en page KPIs + graphiques
# ===========================================================================

def _kpi_card(
    ws: Any,
    row: int,
    col: int,
    label: str,
    value: Any,
    context: str = "",
    value_color: str = _NAVY,
) -> None:
    """Écrit une carte KPI sur 3 lignes : libellé / valeur / contexte.

    La carte utilise des bordures asymétriques (haut seul, côtés seuls, bas seul)
    pour créer un contour propre autour des 3 lignes sans double trait entre elles.

    Args:
        ws: Feuille cible.
        row: Ligne du libellé (les lignes row+1 et row+2 sont occupées).
        col: Colonne de la carte.
        label: Libellé en petit gris en haut.
        value: Valeur principale en grand gras navy (ou couleur custom).
        context: Ligne de contexte en petit italique gris en bas.
        value_color: Couleur de la valeur (pour variations pos/neg).
    """
    thin = _THIN_SIDE

    _w(ws, row,   col, label,   size=9,  color=_GREY_MID,  align="center", bg=_GREY_LITE)
    _w(ws, row+1, col, value,   size=14, color=value_color, bold=True, align="center", bg=_GREY_LITE)
    _w(ws, row+2, col, context, size=8,  color=_GREY_MID,  align="center", bg=_GREY_LITE, italic=True)

    ws.cell(row,   col).border = Border(top=thin,    left=thin, right=thin)
    ws.cell(row+1, col).border = Border(left=thin,   right=thin)
    ws.cell(row+2, col).border = Border(bottom=thin, left=thin, right=thin)

    ws.row_dimensions[row].height   = 13
    ws.row_dimensions[row+1].height = 26
    ws.row_dimensions[row+2].height = 13


def _section_header(ws: Any, row: int, text: str, max_col: int = 6) -> None:
    """Écrit un en-tête de section navy fusionné sur plusieurs colonnes.

    Args:
        ws: Feuille cible.
        row: Numéro de ligne.
        text: Texte de l'en-tête.
        max_col: Dernière colonne de la fusion.
    """
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=max_col,
    )
    _w(ws, row, 1, text, bold=True, size=11, color=_NAVY, align="left")
    ws.row_dimensions[row].height = 18


def _build_dashboard_sheet(
    wb: Any,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    df: pd.DataFrame,
    kpis: DashboardKPIs,
) -> None:
    """Construit la feuille '📊 Dashboard' dans le classeur.

    Layout :
        Ligne 1    : titre principal (fond navy, texte blanc)
        Ligne 2    : sous-titre période + date de génération
        Ligne 3    : vide (respiration)
        Ligne 4    : en-tête section "INDICATEURS CLÉS"
        Lignes 5-7 : 3 cartes KPI (ligne A)
        Ligne 8    : vide
        Lignes 9-11: 3 cartes KPI (ligne B)
        Ligne 12   : vide
        Ligne 13   : en-tête section "ANALYSES GRAPHIQUES"
        Lignes 14+ : 5 graphiques PNG matplotlib (2 par ligne, 3 lignes)

    Args:
        wb: Classeur openpyxl.
        report_months: Agrégation mensuelle [mois, montant] pour la période.
        report_salespeople: Agrégation par commercial [commercial, montant] ou vide.
        df: DataFrame nettoyé complet (pour le calcul du panier moyen).
        kpis: Indicateurs pré-calculés par ``compute_dashboard_kpis``.
    """
    ws = wb.create_sheet("📊 Dashboard")

    # Largeurs de colonnes
    for col_idx, width in _COL_WIDTHS_DB.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    period_label = (
        f"{report_months['mois'].min()} → {report_months['mois'].max()}"
    )

    # -----------------------------------------------------------------------
    # Ligne 1 : Titre principal
    # -----------------------------------------------------------------------
    ws.merge_cells("A1:F1")
    _w(ws, 1, 1,
       f"TABLEAU DE BORD  ·  {SETTINGS.app_name}",
       bold=True, size=18, color=_WHITE, bg=_NAVY, align="center")
    ws.row_dimensions[1].height = 34

    # -----------------------------------------------------------------------
    # Ligne 2 : Sous-titre
    # -----------------------------------------------------------------------
    ws.merge_cells("A2:F2")
    _w(ws, 2, 1,
       f"Période : {period_label}   |   "
       f"Généré le {datetime.today().strftime('%d/%m/%Y')}   |   "
       f"v{SETTINGS.app_version}",
       size=9, color=_GREY_MID, bg=_GREY_LITE, align="center", italic=True)
    ws.row_dimensions[2].height = 15

    # Ligne 3 : respiration
    ws.row_dimensions[3].height = 8

    # -----------------------------------------------------------------------
    # Ligne 4 : En-tête section KPIs
    # -----------------------------------------------------------------------
    _section_header(ws, 4, "INDICATEURS CLÉS")

    # -----------------------------------------------------------------------
    # Lignes 5-7 : KPI Row A (CA total, Panier moyen, Nb transactions)
    # -----------------------------------------------------------------------
    _kpi_card(
        ws, 5, 1,
        "CHIFFRE D'AFFAIRES TOTAL",
        _fmt_eur(kpis.total_revenue),
        f"{kpis.active_months} mois couverts",
    )
    _kpi_card(
        ws, 5, 3,
        "PANIER MOYEN",
        _fmt_eur(kpis.average_basket, decimals=2),
        "par transaction valide",
    )
    _kpi_card(
        ws, 5, 5,
        "NB TRANSACTIONS VALIDES",
        f"{kpis.transaction_count:,}".replace(",", " "),
    )

    # Ligne 8 : respiration
    ws.row_dimensions[8].height = 8

    # -----------------------------------------------------------------------
    # Lignes 9-11 : KPI Row B (Variation M/M, Meilleur mois, Top commercial)
    # -----------------------------------------------------------------------

    # Variation M/M — couleur conditionnelle : vert si ≥ 0, rouge sinon
    if kpis.last_mom_growth_pct is not None:
        mom_val   = f"{kpis.last_mom_growth_pct:+.1f} %"
        mom_color = _GREEN if kpis.last_mom_growth_pct >= 0 else _RED
        mom_ctx   = "dernier mois vs précédent"
    else:
        mom_val   = "— N/A"
        mom_color = _GREY_MID
        mom_ctx   = "données insuffisantes (< 2 mois)"

    _kpi_card(ws, 9, 1, "VARIATION M/M", mom_val, mom_ctx, value_color=mom_color)

    _kpi_card(
        ws, 9, 3,
        "MEILLEUR MOIS",
        kpis.best_month,
        f"{_fmt_eur(kpis.best_month_revenue)}  ·  pire : {kpis.worst_month}",
    )

    if kpis.top_salesperson:
        top_ctx = (
            f"{_fmt_eur(kpis.top_salesperson_revenue or 0)}  ·  "
            f"{kpis.active_salespeople} commerciaux"
        )
        _kpi_card(ws, 9, 5, "TOP COMMERCIAL", kpis.top_salesperson, top_ctx)
    else:
        _kpi_card(ws, 9, 5, "TOP COMMERCIAL", "—", "aucune donnée commerciale")

    # Ligne 12 : respiration
    ws.row_dimensions[12].height = 10

    # -----------------------------------------------------------------------
    # Ligne 13 : En-tête section graphiques
    # -----------------------------------------------------------------------
    _section_header(ws, 13, "ANALYSES GRAPHIQUES")

    # -----------------------------------------------------------------------
    # Graphiques : 2 par ligne (ancres G = col 7 pour les graphiques de droite)
    # Chaque graphique occupe ~18 lignes à hauteur standard.
    # -----------------------------------------------------------------------
    months_sorted = report_months.sort_values("mois").reset_index(drop=True)

    img1 = _chart_ca_mensuel(months_sorted)
    img2 = _chart_ca_cumule(months_sorted)
    img3 = _chart_variation_mom(months_sorted)
    img4 = _chart_top_commerciaux(report_salespeople)
    img5 = _chart_panier_moyen(months_sorted, df)

    ws.add_image(img1, "A14")
    ws.add_image(img2, "H14")

    ws.add_image(img3, "A35")
    if img4 is not None:
        ws.add_image(img4, "H35")
    else:
        # Texte d'absence si pas de colonne commercial dans les sources
        ws.merge_cells("G33:K40")
        _w(ws, 33, 7,
           "Graphique 4 non disponible :\naucune colonne « commercial » dans les fichiers sources.",
           size=10, color=_GREY_MID, align="center", italic=True, bg=_GREY_LITE, wrap=True)

    ws.add_image(img5, "A56")


# ===========================================================================
# FEUILLE RÉCONCILIATION — Rapport d'intégrité formaté
# ===========================================================================

def _build_reconciliation_sheet(
    wb: Any,
    report: ReconciliationReport,
) -> None:
    """Construit la feuille '🔐 Réconciliation' : miroir Excel du checkpoint unifié.

    Rend lisible, directement dans le classeur, le rapport produit par
    ``utils.reconcile_against_manifest`` : une ligne par point de contrôle (les
    5 métriques de l'oracle + les contrôles structurels + l'intégrité référentielle
    des clés étrangères), avec couleur conditionnelle OK / ALERTE. Elle coexiste avec
    le fichier ``log.txt`` (même contenu via ``report.render()``) : les deux supports
    se complètent pour l'audit.

    Args:
        wb: Classeur openpyxl.
        report: Rapport de réconciliation unifié (porte ``checks`` et ``integrity_ok``).
    """
    ws = wb.create_sheet("🔐 Réconciliation")

    # 5 colonnes : Contrôle | Obtenu | Attendu | Statut | Détail
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 50

    # --- Titre ---
    ws.merge_cells("A1:E1")
    _w(ws, 1, 1, "RAPPORT DE RÉCONCILIATION — CHECKPOINT D'INTÉGRITÉ",
       bold=True, size=14, color=_WHITE, bg=_NAVY, align="center")
    ws.row_dimensions[1].height = 30

    # --- Sous-titre ---
    ws.merge_cells("A2:E2")
    _w(ws, 2, 1,
       f"Généré le {datetime.today().strftime('%d/%m/%Y à %H:%M')}  |  "
       f"{SETTINGS.app_name} v{SETTINGS.app_version}",
       size=9, color=_GREY_MID, bg=_GREY_LITE, align="center", italic=True)
    ws.row_dimensions[2].height = 14

    ws.row_dimensions[3].height = 8  # respiration

    # --- Bandeau de résultat global ---
    n_ok = sum(1 for c in report.checks if c.ok)
    n_total = len(report.checks)
    if report.integrity_ok:
        status_text = f"✅  INTÉGRITÉ VALIDÉE — {n_ok}/{n_total} contrôles au vert"
        status_bg, status_col = _GREY_LITE, _GREEN
    else:
        status_text = f"🚨  ALERTE INTÉGRITÉ — {n_ok}/{n_total} contrôles au vert"
        status_bg, status_col = _ALERT_BG, _RED

    ws.merge_cells("A4:E4")
    _w(ws, 4, 1, status_text,
       bold=True, size=13, color=status_col, bg=status_bg, align="center")
    ws.cell(4, 1).border = _THIN_BORDER
    ws.row_dimensions[4].height = 28

    ws.row_dimensions[5].height = 8

    # --- En-tête du tableau des contrôles ---
    header_row = 6
    headers = ["Contrôle", "Obtenu", "Attendu", "Statut", "Détail"]
    for col, label in enumerate(headers, start=1):
        align = "left" if col in (1, 5) else "center"
        _w(ws, header_row, col, label,
           bold=True, size=10, color=_WHITE, bg=_NAVY, align=align)
        ws.cell(header_row, col).border = _THIN_BORDER
    ws.row_dimensions[header_row].height = 18

    # --- Une ligne par point de contrôle, couleur conditionnelle OK / ALERTE ---
    # Un contrôle en alerte est surligné (fond _ALERT_BG) pour ressortir d'un coup
    # d'œil ; les contrôles OK alternent blanc / gris clair pour rester lisibles.
    # Le format monétaire ne s'applique qu'aux lignes dont le libellé porte "€"
    # (le CA réconcilié) ; les autres métriques sont des compteurs entiers.
    money_fmt = '#,##0.00 "€"'
    int_fmt = "#,##0"
    for i, check in enumerate(report.checks):
        r = header_row + 1 + i
        row_bg = _ALERT_BG if not check.ok else (_GREY_LITE if i % 2 == 0 else _WHITE)
        num_fmt = money_fmt if "€" in check.label else int_fmt

        _w(ws, r, 1, check.label, size=10, color=_GREY_DARK, bg=row_bg)
        _w(ws, r, 2, float(check.obtained), size=10, color=_NAVY,
           align="right", bg=row_bg, num_fmt=num_fmt)
        _w(ws, r, 3, float(check.expected), size=10, color=_GREY_DARK,
           align="right", bg=row_bg, num_fmt=num_fmt)
        _w(ws, r, 4, "OK" if check.ok else "ALERTE",
           bold=True, size=10, color=_GREEN if check.ok else _RED,
           align="center", bg=row_bg)
        _w(ws, r, 5, check.detail or "—", size=9,
           color=_GREY_MID, italic=True, bg=row_bg, wrap=True)
        for col in range(1, 6):
            ws.cell(r, col).border = _THIN_BORDER
        ws.row_dimensions[r].height = 16

    # --- Note d'audit (miroir du log) ---
    note_row = header_row + 1 + n_total + 1
    ws.merge_cells(
        start_row=note_row, start_column=1, end_row=note_row, end_column=5
    )
    _w(ws, note_row, 1,
       "ℹ️  Réconciliation contre l'oracle (_manifest_anomalies.json). Ce rapport est "
       "également écrit dans le fichier log.txt du run : les deux sont complémentaires "
       "pour l'audit. Rappel : le CA n'est jamais stocké, il est recalculé en SQL sur "
       "les faits du modèle en étoile.",
       size=8, color=_GREY_MID, italic=True, align="left", wrap=True)
    ws.row_dimensions[note_row].height = 40


# ===========================================================================
# FORMATAGE DES FEUILLES DE DONNÉES
# ===========================================================================

def _format_data_sheet(ws: Any) -> None:
    """Applique le formatage standard aux feuilles de données du classeur enrichi.

    Variante de ``utils._format_excel_sheet`` avec en-têtes navy et alternance
    de fond sur les lignes pour cohérence avec la charte graphique du dashboard.
    Volontairement indépendante de la fonction privée de utils.py pour respecter
    le principe d'encapsulation (dashboard.py est un module autonome).

    Args:
        ws: Feuille openpyxl à formater (modifiée en place).
    """
    from openpyxl.styles import numbers as openpyxl_numbers

    if ws.max_row < 1:
        return

    # En-têtes : fond bleu marine, texte blanc, gras, centré
    for cell in ws[1]:
        cell.font      = _font(bold=True, size=10, color=_WHITE)
        cell.fill      = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ligne 1 figée
    ws.freeze_panes = "A2"

    # Largeurs automatiques avec planchers métier
    headers   = [str(c.value).lower() if c.value else "" for c in ws[1]]
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        header_val = str(col_cells[0].value) if col_cells[0].value else ""
        max_len    = max(
            len(header_val),
            max(
                (len(str(c.value)) for c in col_cells[1:] if c.value is not None),
                default=0,
            ),
        )
        width = min(max(max_len + 4, 12), 40)
        if any(kw in header_val.lower() for kw in ("montant", "euro", "total", "ca")):
            width = max(width, 18)
        ws.column_dimensions[col_letter].width = width

    # Alternance de fond gris très clair / blanc sur les lignes de données
    for idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
        row_bg = _GREY_LITE if idx % 2 == 0 else _WHITE
        for cell in row_cells:
            cell.fill = _fill(row_bg)

    # Formats date et monétaire
    header_idx = {h: i + 1 for i, h in enumerate(headers) if h}
    if "date" in header_idx:
        c = header_idx["date"]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = "DD/MM/YYYY"
    for kw in ("montant en euros", "montant", "total", "ca"):
        if kw in header_idx:
            c = header_idx[kw]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = openpyxl_numbers.FORMAT_CURRENCY_EUR_SIMPLE
            break


# ===========================================================================
# ASSEMBLAGE DU CLASSEUR COMPLET
# ===========================================================================

def _synthese_narrative(s: pd.Series) -> str:
    """Génère un narratif factuel de 3-4 phrases à partir de la ligne de synthèse.

    Le texte est dérivé mécaniquement des signes et seuils (pas d'interprétation
    arbitraire) : sens de la croissance, du mouvement de marge, et de l'écart budget.

    Args:
        s: Ligne unique de la vue ``summary`` (``analytics['summary'].iloc[0]``).

    Returns:
        Le narratif prêt à insérer dans la cellule de lecture.
    """
    parts: list[str] = []

    def fr(x: float, decimals: int = 1) -> str:
        """Formate un nombre à la française (virgule décimale), sans toucher au reste."""
        return f"{x:.{decimals}f}".replace(".", ",")

    # Chiffre d'affaires + croissance annuelle
    if s["ca_yoy_pct"] is not None and not pd.isna(s["ca_yoy_pct"]):
        sens = "en hausse" if s["ca_yoy_pct"] >= 0 else "en baisse"
        parts.append(
            f"Le chiffre d'affaires atteint {_fmt_eur(s['ca'])}, {sens} de "
            f"{fr(abs(s['ca_yoy_pct']))} % sur un an."
        )
    else:
        parts.append(f"Le chiffre d'affaires atteint {_fmt_eur(s['ca'])} sur la période.")

    # Taux de marge + évolution
    tm = s["taux_marge"]
    if s["taux_marge_delta_pts"] is not None and not pd.isna(s["taux_marge_delta_pts"]):
        d = s["taux_marge_delta_pts"]
        mvt = (
            f"en recul de {fr(abs(d))} pts" if d < 0
            else f"en progression de {fr(abs(d))} pts" if d > 0
            else "stable"
        )
        parts.append(f"Le taux de marge s'établit à {fr(tm)} %, {mvt} par rapport à N-1.")
    else:
        parts.append(f"Le taux de marge s'établit à {fr(tm)} %.")

    # Écart au budget (CA et marge)
    if s["ecart_ca_pct"] is not None and not pd.isna(s["ecart_ca_pct"]):
        sens_ca = "sous le budget" if s["ecart_ca_pct"] < 0 else "au-dessus du budget"
        parts.append(
            f"L'activité est {sens_ca} de {fr(abs(s['ecart_ca_pct']))} % sur le CA "
            f"et de {fr(abs(s['ecart_marge_pct']))} % sur la marge."
        )

    # Point d'attention : croissance + érosion de marge simultanées
    if (
        s["ca_yoy_pct"] is not None and not pd.isna(s["ca_yoy_pct"])
        and s["ca_yoy_pct"] > 0
        and s["taux_marge_delta_pts"] is not None
        and not pd.isna(s["taux_marge_delta_pts"])
        and s["taux_marge_delta_pts"] < 0
    ):
        parts.append(
            "Point d'attention : la croissance du CA s'accompagne d'une érosion de la "
            "marge — à investiguer (mix produit, politique de remises)."
        )

    return " ".join(parts)


def _build_synthese_sheet(
    wb: Any,
    analytics: dict[str, pd.DataFrame],
    report: ReconciliationReport,
) -> None:
    """Construit la feuille '🧭 Synthèse DAF' : le one-pager de direction financière.

    Trois rangées de cartes (niveaux / tendances annuelles / budget) avec couleur
    sémantique sur les variances (vert favorable, rouge défavorable), un tampon
    d'intégrité, et un narratif factuel auto-généré. C'est la première feuille du
    classeur — celle qu'un DAF lit en priorité.

    Args:
        wb: Classeur openpyxl.
        analytics: Vues analytiques (sortie de ``utils.build_analytics_views``).
        report: Rapport de réconciliation (pour le tampon d'intégrité).
    """
    ws = wb.create_sheet("🧭 Synthèse DAF")
    for col_idx, width in _COL_WIDTHS_DB.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    s = analytics["summary"].iloc[0]
    monthly = analytics["monthly"]
    period_label = (
        f"{monthly['mois'].min()} → {monthly['mois'].max()}"
        if not monthly.empty
        else "période indéterminée"
    )

    # Ligne 1 : Titre
    ws.merge_cells("A1:F1")
    _w(ws, 1, 1, f"SYNTHÈSE DIRECTION FINANCIÈRE  ·  {SETTINGS.app_name}",
       bold=True, size=16, color=_WHITE, bg=_NAVY, align="center")
    ws.row_dimensions[1].height = 32

    # Ligne 2 : Sous-titre
    ws.merge_cells("A2:F2")
    _w(ws, 2, 1,
       f"Période : {period_label}   |   "
       f"Généré le {datetime.today().strftime('%d/%m/%Y')}   |   v{SETTINGS.app_version}",
       size=9, color=_GREY_MID, bg=_GREY_LITE, align="center", italic=True)
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 8

    # Ligne 4 : Tampon d'intégrité (le chiffre n'a de valeur que s'il est audité)
    if report.integrity_ok:
        stamp, s_col, s_bg = "✅  DONNÉES RÉCONCILIÉES — INTÉGRITÉ VALIDÉE", _GREEN, _GREY_LITE
    else:
        stamp, s_col, s_bg = "🚨  ALERTE INTÉGRITÉ — CHIFFRES À NE PAS DIFFUSER", _RED, _ALERT_BG
    ws.merge_cells("A4:F4")
    _w(ws, 4, 1, stamp, bold=True, size=11, color=s_col, bg=s_bg, align="center")
    ws.cell(4, 1).border = _THIN_BORDER
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 8

    tm_txt = f"{s['taux_marge']:.1f} %".replace(".", ",")

    # Rangée A — niveaux
    _section_header(ws, 6, "INDICATEURS CLÉS")
    _kpi_card(ws, 7, 1, "CHIFFRE D'AFFAIRES", _fmt_eur(s["ca"]), "réel, exercice complet")
    _kpi_card(ws, 7, 3, "MARGE BRUTE", _fmt_eur(s["marge"]), f"taux : {tm_txt}")
    _kpi_card(ws, 7, 5, "TAUX DE MARGE", tm_txt, "marge brute / CA")
    ws.row_dimensions[10].height = 8

    # Rangée B — tendances annuelles (couleur sémantique)
    _kpi_card(ws, 11, 1, "CROISSANCE CA (1 AN)", _fmt_signed_pct(s["ca_yoy_pct"]),
              "réel vs N-1", value_color=_favorable_color(s["ca_yoy_pct"]))
    _kpi_card(ws, 11, 3, "ÉVOLUTION DU TAUX DE MARGE", _fmt_signed_pts(s["taux_marge_delta_pts"]),
              "points de marge vs N-1", value_color=_favorable_color(s["taux_marge_delta_pts"]))
    _kpi_card(ws, 11, 5, "CLIENTS ACTIFS", f"{int(s['clients_actifs']):,}".replace(",", " "),
              "clients avec ventes livrées")
    ws.row_dimensions[14].height = 8

    # Rangée C — budget (couleur sémantique)
    _section_header(ws, 15, "RÉALISÉ VS BUDGET")
    _kpi_card(ws, 16, 1, "ÉCART BUDGET — CA", _fmt_signed_pct(s["ecart_ca_pct"]),
              _fmt_eur(s["ecart_ca"]), value_color=_favorable_color(s["ecart_ca_pct"]))
    _kpi_card(ws, 16, 3, "ÉCART BUDGET — MARGE", _fmt_signed_pct(s["ecart_marge_pct"]),
              _fmt_eur(s["ecart_marge"]), value_color=_favorable_color(s["ecart_marge_pct"]))
    _kpi_card(ws, 16, 5, "CA BUDGÉTÉ", _fmt_eur(s["ca_budget"]), "objectif de la période")
    ws.row_dimensions[19].height = 8

    # Narratif factuel
    _section_header(ws, 20, "LECTURE DE LA DIRECTION FINANCIÈRE")
    ws.merge_cells("A21:F24")
    _w(ws, 21, 1, _synthese_narrative(s), size=10, color=_GREY_DARK,
       align="left", wrap=True, bg=_GREY_LITE)
    ws.cell(21, 1).border = _THIN_BORDER
    for r in range(21, 25):
        ws.row_dimensions[r].height = 18


def _chart_marge_erosion(monthly: pd.DataFrame) -> XLImage:
    """Graphique signature — CA mensuel (barres) vs taux de marge mensuel (courbe).

    À double axe : le CA en barres (axe gauche, €) et le taux de marge en courbe
    (axe droit, %). C'est le graphique qui *raconte* l'histoire d'un DAF : le volume
    se maintient ou progresse pendant que le taux de marge se dégrade. L'axe de droite
    est auto-échelonné (convention des courbes de taux : on ne part pas de zéro, sinon
    la tendance d'un taux serait illisible).

    Args:
        monthly: Vue mensuelle [mois, ca, marge, taux_marge] triée par mois.

    Returns:
        Image PNG openpyxl.
    """
    labels = monthly["mois"].astype(str).tolist()
    ca = monthly["ca"].tolist()
    tm = monthly["taux_marge"].tolist()
    x = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.bar(list(x), ca, color=f"#{_NAVY}", width=0.6, alpha=0.35, zorder=3)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))
    _apply_chart_style(
        ax, "Érosion de la marge : CA (barres) vs taux de marge (courbe)",
        "Mois", "CA mensuel (€)",
    )

    # Axe secondaire : taux de marge en %, courbe rouge (signal de dégradation)
    ax2 = ax.twinx()
    ax2.plot(list(x), tm, color=f"#{_RED}", linewidth=2, marker="o", markersize=3, zorder=5)
    ax2.set_ylabel("Taux de marge (%)", fontsize=8, color=f"#{_RED}")
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(_pct_axis_fmt_unsigned))
    ax2.tick_params(axis="y", labelsize=7.5, labelcolor=f"#{_RED}", length=3, width=0.5)
    ax2.spines["top"].set_visible(False)
    ax2.spines["right"].set_color(f"#{_RED}")
    ax2.spines["right"].set_linewidth(0.5)

    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_marge_par_categorie(by_category: pd.DataFrame, avg_margin: float) -> XLImage:
    """Graphique — taux de marge par catégorie (barres horizontales, vs moyenne).

    Les barres sont colorées en rouge si la catégorie est sous le taux de marge moyen
    (elle dilue la marge globale), en vert sinon. Une ligne pointillée marque la
    moyenne. La part de CA est annotée en bout de barre : on voit alors si une
    catégorie peu margée pèse lourd dans le CA — c'est le cœur du diagnostic « mix ».

    Args:
        by_category: Vue [categorie, ca, marge, taux_marge, part_ca_pct].
        avg_margin: Taux de marge moyen global (référence).

    Returns:
        Image PNG openpyxl.
    """
    d = by_category.sort_values("taux_marge")
    cats = d["categorie"].tolist()
    tm = d["taux_marge"].tolist()
    parts = d["part_ca_pct"].tolist()
    y = range(len(cats))
    colors = [f"#{_RED}" if t < avg_margin else f"#{_GREEN}" for t in tm]

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.barh(list(y), tm, color=colors, height=0.6, zorder=3)
    ax.axvline(avg_margin, color=f"#{_GREY_MID}", linestyle="--", linewidth=1, zorder=4)
    ax.text(avg_margin, len(cats) - 0.3, f" moy. {avg_margin:.0f} %",
            fontsize=6.5, color=f"#{_GREY_MID}", va="top")
    ax.set_yticks(list(y))
    ax.set_yticklabels(cats)
    ax.set_xlim(0, max(tm) * 1.28)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(_pct_axis_fmt_unsigned))
    for i, (t, p) in enumerate(zip(tm, parts)):
        ax.text(t + max(tm) * 0.02, i, f"{p:.0f} % du CA".replace(".", ","),
                va="center", fontsize=6.5, color=f"#{_GREY_DARK}")

    _apply_chart_style(
        ax, "Taux de marge par catégorie (vs moyenne)",
        "Taux de marge (%)", "", grid_axis="x",
    )
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _build_marge_sheet(wb: Any, analytics: dict[str, pd.DataFrame]) -> None:
    """Construit la feuille '💹 Marge & rentabilité' : érosion + analyse par catégorie.

    Deux graphiques empilés (courbe d'érosion, marge par catégorie) suivis d'une table
    détaillée par catégorie avec couleur conditionnelle sur le taux de marge (rouge si
    sous la moyenne globale). La marge est recalculée en amont (jamais stockée).

    Args:
        wb: Classeur openpyxl.
        analytics: Vues analytiques (sortie de ``utils.build_analytics_views``).
    """
    ws = wb.create_sheet("💹 Marge & rentabilité")
    for col, width in {"A": 24, "B": 16, "C": 16, "D": 16, "E": 14}.items():
        ws.column_dimensions[col].width = width

    monthly = analytics["monthly"]
    by_cat = analytics["by_category"]
    avg_margin = float(analytics["summary"].iloc[0]["taux_marge"])

    # Titre + sous-titre
    ws.merge_cells("A1:F1")
    _w(ws, 1, 1, "MARGE & RENTABILITÉ", bold=True, size=16,
       color=_WHITE, bg=_NAVY, align="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:F2")
    _w(ws, 2, 1,
       f"Taux de marge global : {avg_margin:.1f} %".replace(".", ",")
       + f"   |   Marge brute : {_fmt_eur(monthly['marge'].sum())}",
       size=9, color=_GREY_MID, bg=_GREY_LITE, align="center", italic=True)
    ws.row_dimensions[2].height = 15

    # Graphique 1 : érosion
    _section_header(ws, 4, "ÉROSION DE LA MARGE DANS LE TEMPS")
    ws.add_image(_chart_marge_erosion(monthly), "A5")

    # Graphique 2 : marge par catégorie
    _section_header(ws, 27, "RENTABILITÉ PAR CATÉGORIE")
    ws.add_image(_chart_marge_par_categorie(by_cat, avg_margin), "A28")

    # Table détaillée par catégorie
    _section_header(ws, 50, "DÉTAIL PAR CATÉGORIE")
    headers = ["Catégorie", "CA", "Marge brute", "Taux de marge", "Part de CA"]
    for c, label in enumerate(headers, start=1):
        align = "left" if c == 1 else "center"
        _w(ws, 51, c, label, bold=True, size=10, color=_WHITE, bg=_NAVY, align=align)
        ws.cell(51, c).border = _THIN_BORDER
    money_fmt = '#,##0 "€"'
    pct_fmt = '0.0 "%"'
    d = by_cat.sort_values("ca", ascending=False).reset_index(drop=True)
    for i, row in d.iterrows():
        r = 52 + i
        bg = _GREY_LITE if i % 2 == 0 else _WHITE
        # Couleur conditionnelle : taux sous la moyenne globale = dilutif (rouge)
        tm_color = _RED if row["taux_marge"] < avg_margin else _GREEN
        _w(ws, r, 1, row["categorie"], size=10, color=_GREY_DARK, bg=bg)
        _w(ws, r, 2, float(row["ca"]), size=10, color=_NAVY, align="right",
           bg=bg, num_fmt=money_fmt)
        _w(ws, r, 3, float(row["marge"]), size=10, color=_GREY_DARK, align="right",
           bg=bg, num_fmt=money_fmt)
        _w(ws, r, 4, float(row["taux_marge"]), size=10, color=tm_color, bold=True,
           align="right", bg=bg, num_fmt=pct_fmt)
        _w(ws, r, 5, float(row["part_ca_pct"]), size=10, color=_GREY_DARK,
           align="right", bg=bg, num_fmt=pct_fmt)
        for c in range(1, 6):
            ws.cell(r, c).border = _THIN_BORDER

    # Graphique : évolution du mix produit — explique visuellement l'érosion
    # (une catégorie peu margée qui gagne en part dilue la marge globale).
    _section_header(ws, 59, "ÉVOLUTION DU MIX PRODUIT")
    ws.add_image(_chart_mix_categoriel(analytics["category_mix"]), "A60")

    # Graphique : impact des remises sur la marge (fuite de marge)
    _section_header(ws, 83, "REMISES & FUITE DE MARGE")
    ws.add_image(_chart_remises(analytics["by_discount_band"]), "A84")


def _chart_remises(by_discount_band: pd.DataFrame) -> XLImage:
    """Graphique — taux de marge par tranche de remise, avec part de CA annotée.

    Met en évidence la « fuite de marge » : à mesure que la remise courante augmente,
    le taux de marge se dégrade. La part de CA annotée montre où se concentre le
    volume — donc où une meilleure discipline de remise rapporterait le plus.

    Args:
        by_discount_band: Vue [tranche, lignes, ca, taux_marge, part_ca_pct] ordonnée.

    Returns:
        Image PNG openpyxl.
    """
    bands = by_discount_band["tranche"].tolist()
    tm = by_discount_band["taux_marge"].tolist()
    parts = by_discount_band["part_ca_pct"].tolist()
    x = range(len(bands))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")
    ax.bar(list(x), tm, color=f"#{_NAVY}", width=0.55, zorder=3)
    ax.set_xticks(list(x))
    ax.set_xticklabels(bands)
    ax.set_ylim(0, max(tm) * 1.25)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_pct_axis_fmt_unsigned))
    for i, (t, p) in enumerate(zip(tm, parts)):
        ax.text(i, t + max(tm) * 0.03, f"{p:.0f} % du CA".replace(".", ","),
                ha="center", fontsize=7, color=f"#{_GREY_DARK}")
    _apply_chart_style(ax, "Taux de marge par tranche de remise",
                       "Tranche de remise", "Taux de marge (%)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_budget_mensuel(budget_monthly: pd.DataFrame) -> XLImage:
    """Graphique — CA mensuel réel (barres) vs budget (ligne cible).

    Les barres donnent le réalisé, la ligne pointillée la cible budgétaire : un mois
    dont la barre passe sous la ligne est en retard sur le plan. Lecture immédiate de
    la tenue du budget mois par mois.

    Args:
        budget_monthly: Vue [mois, ca_reel, ca_budget, ...] triée par mois.

    Returns:
        Image PNG openpyxl.
    """
    labels = budget_monthly["mois"].astype(str).tolist()
    reel = budget_monthly["ca_reel"].tolist()
    budget = budget_monthly["ca_budget"].tolist()
    x = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.bar(list(x), reel, color=f"#{_NAVY}", width=0.6, zorder=3, label="Réel")
    ax.plot(list(x), budget, color=f"#{_RED}", linewidth=1.5, linestyle="--",
            marker="o", markersize=2.5, zorder=5, label="Budget")
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))
    ax.legend(fontsize=7, loc="upper left", framealpha=0.9)

    _apply_chart_style(ax, "CA mensuel : réel vs budget", "Mois", "Montant (€)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_budget_ecart_region(budget_by_region: pd.DataFrame) -> XLImage:
    """Graphique — écart au budget (CA, %) par région (barres horizontales signées).

    Barres vertes (au-dessus du plan) ou rouges (sous le plan), triées : la région la
    plus en retard ressort immédiatement. Ici le signe est porteur de sens (c'est un
    écart, pas un niveau), d'où l'axe en pourcentage signé.

    Args:
        budget_by_region: Vue [region, ecart_ca_pct, ...].

    Returns:
        Image PNG openpyxl.
    """
    d = budget_by_region.sort_values("ecart_ca_pct")
    regions = d["region"].tolist()
    ecarts = d["ecart_ca_pct"].tolist()
    y = range(len(regions))
    colors = [f"#{_GREEN}" if e >= 0 else f"#{_RED}" for e in ecarts]

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.barh(list(y), ecarts, color=colors, height=0.6, zorder=3)
    ax.axvline(0, color=f"#{_GREY_MID}", linewidth=0.8, zorder=4)
    ax.set_yticks(list(y))
    ax.set_yticklabels(regions)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(_pct_axis_fmt))

    _apply_chart_style(
        ax, "Écart au budget (CA) par région", "Écart (%)", "", grid_axis="x",
    )
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _build_budget_sheet(wb: Any, analytics: dict[str, pd.DataFrame]) -> None:
    """Construit la feuille '🎯 Réalisé vs Budget' : suivi mensuel + détail régional.

    Deux graphiques (CA réel vs budget par mois, écart par région) suivis d'une table
    régionale couvrant CA et marge, avec couleur conditionnelle sur les écarts (rouge
    si sous le budget). Convention : écart = réel - budget.

    Args:
        wb: Classeur openpyxl.
        analytics: Vues analytiques (sortie de ``utils.build_analytics_views``).
    """
    ws = wb.create_sheet("🎯 Réalisé vs Budget")
    widths = {"A": 20, "B": 15, "C": 15, "D": 13, "E": 15, "F": 15, "G": 13}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    bm = analytics["budget_monthly"]
    br = analytics["budget_by_region"]
    s = analytics["summary"].iloc[0]

    # Titre + sous-titre (écart global)
    ws.merge_cells("A1:G1")
    _w(ws, 1, 1, "RÉALISÉ VS BUDGET", bold=True, size=16,
       color=_WHITE, bg=_NAVY, align="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:G2")
    _w(ws, 2, 1,
       f"Écart global : {_fmt_signed_pct(s['ecart_ca_pct'])} sur le CA   |   "
       f"{_fmt_signed_pct(s['ecart_marge_pct'])} sur la marge",
       size=9, color=_GREY_MID, bg=_GREY_LITE, align="center", italic=True)
    ws.row_dimensions[2].height = 15

    # Graphiques
    _section_header(ws, 4, "CA MENSUEL : RÉEL VS BUDGET", max_col=7)
    ws.add_image(_chart_budget_mensuel(bm), "A5")

    _section_header(ws, 27, "ÉCART AU BUDGET PAR RÉGION", max_col=7)
    ws.add_image(_chart_budget_ecart_region(br), "A28")

    # Table détaillée par région (CA + marge)
    _section_header(ws, 50, "DÉTAIL PAR RÉGION", max_col=7)
    headers = [
        "Région", "CA réel", "CA budget", "Écart CA",
        "Marge réelle", "Marge budget", "Écart marge",
    ]
    for c, label in enumerate(headers, start=1):
        align = "left" if c == 1 else "center"
        _w(ws, 51, c, label, bold=True, size=9, color=_WHITE, bg=_NAVY,
           align=align, wrap=True)
        ws.cell(51, c).border = _THIN_BORDER
    ws.row_dimensions[51].height = 26

    money_fmt = '#,##0 "€"'
    pct_signed = '+0.0" %";-0.0" %"'
    d = br.sort_values("ca_reel", ascending=False).reset_index(drop=True)
    for i, row in d.iterrows():
        r = 52 + i
        bg = _GREY_LITE if i % 2 == 0 else _WHITE
        ca_col = _RED if row["ecart_ca_pct"] < 0 else _GREEN
        mg_col = _RED if row["ecart_marge_pct"] < 0 else _GREEN
        _w(ws, r, 1, row["region"], size=10, color=_GREY_DARK, bg=bg)
        _w(ws, r, 2, float(row["ca_reel"]), size=10, color=_NAVY, align="right",
           bg=bg, num_fmt=money_fmt)
        _w(ws, r, 3, float(row["ca_budget"]), size=10, color=_GREY_DARK, align="right",
           bg=bg, num_fmt=money_fmt)
        _w(ws, r, 4, float(row["ecart_ca_pct"]), size=10, color=ca_col, bold=True,
           align="right", bg=bg, num_fmt=pct_signed)
        _w(ws, r, 5, float(row["marge_reel"]), size=10, color=_NAVY, align="right",
           bg=bg, num_fmt=money_fmt)
        _w(ws, r, 6, float(row["marge_budget"]), size=10, color=_GREY_DARK,
           align="right", bg=bg, num_fmt=money_fmt)
        _w(ws, r, 7, float(row["ecart_marge_pct"]), size=10, color=mg_col, bold=True,
           align="right", bg=bg, num_fmt=pct_signed)
        for c in range(1, 8):
            ws.cell(r, c).border = _THIN_BORDER


def _chart_mix_categoriel(category_mix: pd.DataFrame) -> XLImage:
    """Graphique — évolution du mix produit (barres empilées de part de CA par an).

    Chaque barre (une par année) somme à 100 % ; les segments sont les catégories.
    On *voit* le glissement du mix (une catégorie peu margée qui gagne en part dilue
    mécaniquement la marge globale). C'est l'explication visuelle de l'érosion.

    Args:
        category_mix: Vue [categorie, part_AAAA, ...] (une colonne de part par année).

    Returns:
        Image PNG openpyxl.
    """
    palette = ["1F3864", "2E75B6", "8FAADC", "375623", "7F9C5A", "C9A227", "9C5A2E"]
    year_cols = [c for c in category_mix.columns if c.startswith("part_")]
    years = [c.replace("part_", "") for c in year_cols]
    cats = category_mix["categorie"].tolist()
    x = range(len(years))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    bottom = [0.0] * len(years)
    for i, cat in enumerate(cats):
        vals = [
            float(category_mix.loc[category_mix["categorie"] == cat, yc].fillna(0).iloc[0])
            for yc in year_cols
        ]
        ax.bar(list(x), vals, bottom=bottom, width=0.5, zorder=3,
               color=f"#{palette[i % len(palette)]}", label=cat)
        bottom = [b + v for b, v in zip(bottom, vals)]

    ax.set_xticks(list(x))
    ax.set_xticklabels(years)
    ax.set_ylim(0, 100)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_pct_axis_fmt_unsigned))
    _apply_chart_style(ax, "Évolution du mix produit (part de CA)", "Année", "Part de CA (%)")
    ax.legend(fontsize=6.5, loc="upper left", bbox_to_anchor=(1.01, 1.0), framealpha=0.9)
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_segment_poids(by_segment: pd.DataFrame) -> XLImage:
    """Graphique — poids de chaque segment dans le CA (barres horizontales).

    Annoté du nombre de clients : on voit qu'un segment peut peser lourd en CA tout
    en reposant sur peu de clients (concentration), ou l'inverse (diversification).

    Args:
        by_segment: Vue [segment, ca, part_ca_pct, nb_clients, ...].

    Returns:
        Image PNG openpyxl.
    """
    d = by_segment.sort_values("part_ca_pct")
    segs = d["segment"].tolist()
    parts = d["part_ca_pct"].tolist()
    nbs = d["nb_clients"].tolist()
    y = range(len(segs))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")
    ax.barh(list(y), parts, color=f"#{_NAVY}", height=0.55, zorder=3)
    ax.set_yticks(list(y))
    ax.set_yticklabels(segs)
    ax.set_xlim(0, max(parts) * 1.25)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(_pct_axis_fmt_unsigned))
    for i, (p, n) in enumerate(zip(parts, nbs)):
        ax.text(p + max(parts) * 0.02, i, f"{n} clients", va="center",
                fontsize=7, color=f"#{_GREY_DARK}")
    _apply_chart_style(ax, "Part du CA par segment client", "Part de CA (%)", "",
                       grid_axis="x")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_client_pareto(client_pareto: pd.DataFrame) -> XLImage:
    """Graphique — courbe de Pareto du CA client (concentration / diversification).

    Trace le CA cumulé (%) en fonction du nombre de clients. Une courbe qui grimpe
    vite = forte dépendance à quelques clients ; une courbe progressive = base
    diversifiée. Repère à 80 % du CA avec le nombre de clients correspondant.

    Args:
        client_pareto: Vue [rang, ca, ca_cumule_pct] triée par CA décroissant.

    Returns:
        Image PNG openpyxl.
    """
    rangs = client_pareto["rang"].tolist()
    cumul = client_pareto["ca_cumule_pct"].tolist()
    n_total = len(rangs)
    n80 = int((client_pareto["ca_cumule_pct"] <= 80).sum()) or 1

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")
    ax.plot(rangs, cumul, color=f"#{_NAVY}", linewidth=2, zorder=4)
    ax.fill_between(rangs, cumul, color=f"#{_NAVY}", alpha=0.08, zorder=3)
    ax.axhline(80, color=f"#{_GREY_MID}", linestyle="--", linewidth=0.8, zorder=2)
    ax.axvline(n80, color=f"#{_GREY_MID}", linestyle="--", linewidth=0.8, zorder=2)
    ax.text(n80, 40, f" {n80} clients ({100 * n80 / n_total:.0f} %)\n pour 80 % du CA",
            fontsize=7, color=f"#{_GREY_DARK}", va="center")
    ax.set_ylim(0, 100)
    ax.set_xlim(0, n_total)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_pct_axis_fmt_unsigned))
    _apply_chart_style(ax, "Concentration du CA par client (Pareto)",
                       "Nombre de clients (du plus gros au plus petit)", "CA cumulé (%)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _build_salesperson_sheet(wb: Any, analytics: dict[str, pd.DataFrame]) -> None:
    """Construit la feuille '👤 Par commercial' enrichie : CA, marge et taux par tête.

    Remplace l'ancien export brut (CA seul) : un commercial qui « fait du volume »
    à coups de remises a un CA élevé mais un taux de marge faible. La couleur
    conditionnelle sur le taux (rouge si sous la moyenne) rend cela visible.

    Args:
        wb: Classeur openpyxl.
        analytics: Vues analytiques (contient ``by_salesperson`` et ``summary``).
    """
    ws = wb.create_sheet("👤 Par commercial")
    for col, width in {"A": 22, "B": 16, "C": 16, "D": 15, "E": 13}.items():
        ws.column_dimensions[col].width = width

    sp = analytics["by_salesperson"].copy()
    avg_margin = float(analytics["summary"].iloc[0]["taux_marge"])

    ws.merge_cells("A1:E1")
    _w(ws, 1, 1, "PERFORMANCE PAR COMMERCIAL", bold=True, size=14,
       color=_WHITE, bg=_NAVY, align="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:E2")
    _w(ws, 2, 1, "CA, marge brute et taux de marge par commercial "
       f"(taux moyen : {avg_margin:.1f} %".replace(".", ",") + ")",
       size=9, color=_GREY_MID, bg=_GREY_LITE, align="center", italic=True)
    ws.row_dimensions[2].height = 14

    headers = ["Commercial", "CA", "Marge brute", "Taux de marge", "Part de CA"]
    for c, label in enumerate(headers, start=1):
        align = "left" if c == 1 else "center"
        _w(ws, 4, c, label, bold=True, size=10, color=_WHITE, bg=_NAVY, align=align)
        ws.cell(4, c).border = _THIN_BORDER

    money_fmt = '#,##0 "€"'
    pct_fmt = '0.0 "%"'
    for i, row in sp.reset_index(drop=True).iterrows():
        r = 5 + i
        bg = _GREY_LITE if i % 2 == 0 else _WHITE
        tm_color = _RED if row["taux_marge"] < avg_margin else _GREEN
        _w(ws, r, 1, row["commercial"], size=10, color=_GREY_DARK, bg=bg)
        _w(ws, r, 2, float(row["ca"]), size=10, color=_NAVY, align="right",
           bg=bg, num_fmt=money_fmt)
        _w(ws, r, 3, float(row["marge"]), size=10, color=_GREY_DARK, align="right",
           bg=bg, num_fmt=money_fmt)
        _w(ws, r, 4, float(row["taux_marge"]), size=10, color=tm_color, bold=True,
           align="right", bg=bg, num_fmt=pct_fmt)
        _w(ws, r, 5, float(row["part_ca_pct"]), size=10, color=_GREY_DARK,
           align="right", bg=bg, num_fmt=pct_fmt)
        for c in range(1, 6):
            ws.cell(r, c).border = _THIN_BORDER


def _build_clients_sheet(wb: Any, analytics: dict[str, pd.DataFrame]) -> None:
    """Construit la feuille '👥 Clients & segments' : poids segments + concentration.

    Cadrage : la concentration se lit d'abord au niveau segment (les Grands
    Comptes pèsent lourd avec peu de clients), tandis que la base de clients prise
    individuellement est, elle, diversifiée — un atout (faible dépendance). La courbe
    de Pareto matérialise cette diversification.

    Args:
        wb: Classeur openpyxl.
        analytics: Vues analytiques (``by_segment``, ``client_pareto``, ``client_top``).
    """
    ws = wb.create_sheet("👥 Clients & segments")
    for col, width in {"A": 34, "B": 18, "C": 13, "D": 13, "E": 13}.items():
        ws.column_dimensions[col].width = width

    by_segment = analytics["by_segment"]
    pareto = analytics["client_pareto"]
    top = analytics["client_top"]

    n_total = len(pareto)
    n80 = int((pareto["ca_cumule_pct"] <= 80).sum()) or 1
    top_seg = by_segment.iloc[0]

    ws.merge_cells("A1:E1")
    _w(ws, 1, 1, "CLIENTS & SEGMENTS", bold=True, size=14,
       color=_WHITE, bg=_NAVY, align="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:E2")
    _w(ws, 2, 1,
       f"{top_seg['segment']} : {top_seg['part_ca_pct']:.0f} % du CA pour "
       f"{int(top_seg['nb_clients'])} clients   |   base diversifiée : "
       f"{n80} clients ({100 * n80 / n_total:.0f} %) pour 80 % du CA",
       size=9, color=_GREY_MID, bg=_GREY_LITE, align="center", italic=True)
    ws.row_dimensions[2].height = 14

    # Graphiques
    _section_header(ws, 4, "RÉPARTITION DU CA PAR SEGMENT", max_col=5)
    ws.add_image(_chart_segment_poids(by_segment), "A5")
    _section_header(ws, 27, "CONCENTRATION DU CA PAR CLIENT (PARETO)", max_col=5)
    ws.add_image(_chart_client_pareto(pareto), "A28")

    # Table segments
    _section_header(ws, 50, "DÉTAIL PAR SEGMENT", max_col=5)
    seg_headers = ["Segment", "CA", "Part de CA", "Nb clients", "Taux de marge"]
    for c, label in enumerate(seg_headers, start=1):
        align = "left" if c == 1 else "center"
        _w(ws, 51, c, label, bold=True, size=10, color=_WHITE, bg=_NAVY, align=align)
        ws.cell(51, c).border = _THIN_BORDER
    money_fmt = '#,##0 "€"'
    pct_fmt = '0.0 "%"'
    for i, row in by_segment.reset_index(drop=True).iterrows():
        r = 52 + i
        bg = _GREY_LITE if i % 2 == 0 else _WHITE
        _w(ws, r, 1, row["segment"], size=10, color=_GREY_DARK, bg=bg)
        _w(ws, r, 2, float(row["ca"]), size=10, color=_NAVY, align="right",
           bg=bg, num_fmt=money_fmt)
        _w(ws, r, 3, float(row["part_ca_pct"]), size=10, color=_GREY_DARK,
           align="right", bg=bg, num_fmt=pct_fmt)
        _w(ws, r, 4, int(row["nb_clients"]), size=10, color=_GREY_DARK,
           align="right", bg=bg, num_fmt="#,##0")
        _w(ws, r, 5, float(row["taux_marge"]), size=10, color=_GREY_DARK,
           align="right", bg=bg, num_fmt=pct_fmt)
        for c in range(1, 6):
            ws.cell(r, c).border = _THIN_BORDER

    # Table top clients
    base = 52 + len(by_segment) + 1
    _section_header(ws, base, "TOP 10 CLIENTS", max_col=5)
    top_headers = ["Client", "Segment", "CA", "Part de CA", "Taux de marge"]
    for c, label in enumerate(top_headers, start=1):
        align = "left" if c in (1, 2) else "center"
        _w(ws, base + 1, c, label, bold=True, size=10, color=_WHITE, bg=_NAVY, align=align)
        ws.cell(base + 1, c).border = _THIN_BORDER
    for i, row in top.head(10).reset_index(drop=True).iterrows():
        r = base + 2 + i
        bg = _GREY_LITE if i % 2 == 0 else _WHITE
        _w(ws, r, 1, row["raison_sociale"], size=10, color=_GREY_DARK, bg=bg)
        _w(ws, r, 2, row["segment"], size=10, color=_GREY_MID, bg=bg)
        _w(ws, r, 3, float(row["ca"]), size=10, color=_NAVY, align="right",
           bg=bg, num_fmt=money_fmt)
        _w(ws, r, 4, float(row["part_ca_pct"]), size=10, color=_GREY_DARK,
           align="right", bg=bg, num_fmt=pct_fmt)
        _w(ws, r, 5, float(row["taux_marge"]), size=10, color=_GREY_DARK,
           align="right", bg=bg, num_fmt=pct_fmt)
        for c in range(1, 6):
            ws.cell(r, c).border = _THIN_BORDER


def _assemble_workbook(
    df: pd.DataFrame,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    reconciliation_report: ReconciliationReport,
    analytics: dict[str, pd.DataFrame],
) -> Any:
    """Assemble le classeur Excel enrichi avec toutes les feuilles dans l'ordre.

    Ordre final des feuilles (visible pour l'utilisateur) :
        1. 🧭 Synthèse DAF
        2. 💹 Marge & rentabilité
        3. 🎯 Réalisé vs Budget
        4. 👥 Clients & segments
        5. 📊 Dashboard
        6. 📅 Par mois
        7. 👤 Par commercial
        8. 🔍 Annexe (données brutes)
        9. 🔐 Réconciliation

    Args:
        df: DataFrame nettoyé.
        report_months: Agrégation mensuelle [mois, montant].
        report_salespeople: Agrégation par commercial [commercial, montant].
        reconciliation_report: Rapport issu de ``utils.reconcile_against_manifest()``.
        analytics: Vues analytiques DAF (``utils.build_analytics_views``).

    Returns:
        Classeur openpyxl complet, ordonné et formaté, prêt à être sauvegardé.
    """
    # Copies pour l'export : renommage de la colonne interne "montant"
    # en "montant en euros" pour la lisibilité dans les feuilles de données.
    months_xp = report_months.copy().rename(columns={"montant": "montant en euros"})

    # Création initiale via pandas ExcelWriter (gère la sérialisation des DataFrames).
    # « Par commercial » n'est plus un export brut : il est enrichi (CA + marge + taux)
    # par _build_salesperson_sheet à partir des vues analytiques.
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        months_xp.to_excel(writer, sheet_name="📅 Par mois",            index=False)
        df.to_excel(       writer, sheet_name="🔍 Annexe (données brutes)", index=False)
    buf.seek(0)
    wb = load_workbook(buf)

    # Formatage des feuilles de données
    for name in ["📅 Par mois", "🔍 Annexe (données brutes)"]:
        if name in wb.sheetnames:
            _format_data_sheet(wb[name])
    # L'annexe est un export d'audit volumineux : on la rend filtrable (autofilter)
    # plutôt que de la donner brute — un DAF ne déroule pas 50 000+ lignes à la main.
    annexe = wb["🔍 Annexe (données brutes)"]
    annexe.auto_filter.ref = annexe.dimensions

    # Calcul des KPIs
    kpis = compute_dashboard_kpis(df, report_months, report_salespeople)

    # Ajout des nouvelles feuilles
    _build_synthese_sheet(wb, analytics, reconciliation_report)
    _build_marge_sheet(wb, analytics)
    _build_budget_sheet(wb, analytics)
    _build_clients_sheet(wb, analytics)
    _build_dashboard_sheet(wb, report_months, report_salespeople, df, kpis)
    _build_salesperson_sheet(wb, analytics)
    _build_reconciliation_sheet(wb, reconciliation_report)

    # Réordonnancement : ordre absolu voulu. Manipulation de wb._sheets (liste interne
    # openpyxl 3.x) : l'API publique move_sheet() ne permet pas d'ordre absolu.
    target_order = [
        "🧭 Synthèse DAF",
        "💹 Marge & rentabilité",
        "🎯 Réalisé vs Budget",
        "👥 Clients & segments",
        "📊 Dashboard",
        "📅 Par mois",
        "👤 Par commercial",
        "🔍 Annexe (données brutes)",
        "🔐 Réconciliation",
    ]
    sheet_map  = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [sheet_map[name] for name in target_order if name in sheet_map]

    # Mise en page impression : paysage + ajustement à la largeur d'une page.
    # Les feuilles larges (réconciliation, budget, clients) ne se scindent plus
    # latéralement à l'impression ; les lignes continuent sur plusieurs pages.
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    return wb


# ===========================================================================
# FONCTIONS PUBLIQUES — API Streamlit et Batch
# ===========================================================================

def build_excel_bytes_with_dashboard(
    df: pd.DataFrame,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    reconciliation_report: ReconciliationReport,
    analytics: dict[str, pd.DataFrame],
) -> bytes:
    """Génère le classeur Excel enrichi en mémoire (mode Streamlit).

    Toutes les feuilles sont créées, formatées et ordonnées en un seul passage.

    Args:
        df: DataFrame nettoyé (colonnes : date, montant, mois, commercial optionnel).
        report_months: Agrégation mensuelle [mois, montant].
        report_salespeople: Agrégation par commercial [commercial, montant] ou vide.
        reconciliation_report: Rapport issu de ``utils.reconcile_against_manifest()``.
        analytics: Vues analytiques DAF (``utils.build_analytics_views``).

    Returns:
        Bytes du fichier ``.xlsx`` complet, prêt à être proposé en téléchargement.
    """
    wb  = _assemble_workbook(
        df, report_months, report_salespeople, reconciliation_report, analytics
    )
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    logging.info(
        "Classeur Excel enrichi généré en mémoire "
        "(synthèse + dashboard + réconciliation + données formatées)."
    )
    return out.getvalue()


def export_excel_with_dashboard(
    df: pd.DataFrame,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    reconciliation_report: ReconciliationReport,
    filepath: Path,
    analytics: dict[str, pd.DataFrame],
) -> Path:
    """Exporte le classeur Excel enrichi sur le disque (mode Batch).

    Crée le répertoire parent si nécessaire.

    Args:
        df: DataFrame nettoyé.
        report_months: Agrégation mensuelle.
        report_salespeople: Agrégation par commercial.
        reconciliation_report: Rapport de réconciliation.
        filepath: Chemin de destination (fichier ``.xlsx``).
        analytics: Vues analytiques DAF (``utils.build_analytics_views``).

    Returns:
        Path du fichier créé.

    Raises:
        Exception: Propagée si l'écriture sur disque échoue (disque plein, droits…).
    """
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    try:
        wb = _assemble_workbook(
            df, report_months, report_salespeople, reconciliation_report, analytics
        )
        wb.save(str(filepath))
        logging.info("Classeur Excel enrichi exporté : %s", filepath)
        return Path(filepath)
    except Exception as exc:
        logging.error("Erreur export classeur enrichi : %s | %s", filepath, exc)
        raise