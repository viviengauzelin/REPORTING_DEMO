"""
generate_demo_data.py - Générateur de données de démonstration (distributeur B2B CHR).

Simule le système d'information d'un distributeur B2B pour la restauration
(cafés, hôtels, restaurants). Produit 5 fichiers sources « sales » + 1 manifeste
de vérité (oracle indépendant) destiné à la réconciliation et aux tests pytest.

Principes (cf. spec figée) :
- Signaux métier (mix, churn, sous-budget…) ≠ anomalies qualité (formats, doublons…).
- La saleté est majoritairement *systématique* (propriété d'un export), pas du bruit aléatoire.
- L'oracle (manifeste) est calculé sur les valeurs PROPRES, avant tout salissage.

Ce module est déterministe : une même graine produit exactement les mêmes fichiers.

Usage :
    python generate_demo_data.py
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

import numpy as np
import pandas as pd

# ===========================================================================
# CONSTANTES GLOBALES (source de vérité du paramétrage - cf. spec §2)
# ===========================================================================

SEED: int = 42  # Reproductibilité totale : même graine => mêmes fichiers + manifeste.

# Période couverte : 24 mois pour permettre l'analyse en glissement annuel (YoY).
PERIOD_START: str = "2024-01"
PERIOD_END: str = "2025-12"

# Volumétrie cible (cf. spec §2). Assez grande pour être non triviale, assez
# légère pour tenir dans DuckDB/Postgres/Power BI et sur un dépôt Git.
N_CUSTOMERS: int = 500
N_PRODUCTS: int = 120
N_SALESPEOPLE: int = 10
N_REGIONS: int = 5

# Croissance de base d'une année sur l'autre : le CA 2025 part ~8 % au-dessus de
# 2024 afin que l'analyse YoY révèle une tendance haussière crédible.
YOY_GROWTH: float = 1.08

# --- Référentiel géographique : 5 régions, 2 commerciaux chacune (cf. spec §3.1) ---
# La région Sud (R04) portera le signal "sous-budget persistant" (injecté côté budget).
REGIONS: dict[str, str] = {
    "R01": "Île-de-France",
    "R02": "Nord",
    "R03": "Est",
    "R04": "Sud",
    "R05": "Ouest",
}
UNDERBUDGET_REGION_CODE: str = "R04"  # Sud

# Villes par région (pour des données clients crédibles).
CITIES_BY_REGION: dict[str, tuple[str, ...]] = {
    "R01": ("Paris", "Versailles", "Nanterre", "Boulogne-Billancourt"),
    "R02": ("Lille", "Roubaix", "Dunkerque", "Valenciennes"),
    "R03": ("Strasbourg", "Metz", "Nancy", "Mulhouse"),
    "R04": ("Marseille", "Nice", "Toulon", "Aix-en-Provence"),
    "R05": ("Nantes", "Rennes", "Brest", "Angers"),
}

# --- Signaux commerciaux (appliqués lors de la génération des faits) ---
# Le code des commerciaux "en montée" et "en déclin" est figé ici pour que le
# signal soit reproductible et documenté.
RISING_SALESPERSON_CODE: str = "C09"      # recrue récente, montée en charge progressive
RISING_SALESPERSON_HIRE: str = "2024-09-01"
DECLINING_SALESPERSON_CODE: str = "C01"   # activité en repli sur 24 mois

# Prénoms / noms pour fabriquer des identités commerciales lisibles.
_FIRST_NAMES: tuple[str, ...] = (
    "Alice", "Bruno", "Chloé", "David", "Emma",
    "Farid", "Gaëlle", "Hugo", "Inès", "Julien",
)
_LAST_NAMES: tuple[str, ...] = (
    "Martin", "Leroy", "Dupont", "Moreau", "Lefebvre",
    "Garnier", "Rousseau", "Blanc", "Faure", "Mercier",
)

# --- Catalogue produit : 6 catégories aux profils de marge volontairement opposés ---
# C'est ce contraste qui rend l'effet mix lisible (le coeur de l'analyse) :
# les consommables sont à marge élevée et petit ticket, l'équipement à marge
# faible et gros ticket. Si le mix glisse vers l'équipement, le CA monte mais le
# taux de marge global se dégrade.
@dataclass(frozen=True)
class CategorySpec:
    """Profil économique d'une catégorie produit.

    Attributes:
        share: Part de la catégorie dans le catalogue (somme = 1.0).
        cost_min: Coût unitaire minimal (€).
        cost_max: Coût unitaire maximal (€).
        margin_min: Taux de marge cible minimal (ex: 0.45 = 45 %).
        margin_max: Taux de marge cible maximal.
        qty_min: Quantité minimale par ligne de commande.
        qty_max: Quantité maximale par ligne de commande.
        is_equipment: True si la catégorie appartient à la famille "équipement"
            (sert au calcul du signal de glissement de mix).
        sub_categories: Sous-catégories possibles.
    """

    share: float
    cost_min: float
    cost_max: float
    margin_min: float
    margin_max: float
    qty_min: int
    qty_max: int
    is_equipment: bool
    sub_categories: tuple[str, ...]


CATEGORIES: dict[str, CategorySpec] = {
    "Consommables": CategorySpec(
        share=0.30, cost_min=1, cost_max=12, margin_min=0.45, margin_max=0.60,
        qty_min=5, qty_max=80, is_equipment=False,
        sub_categories=("Vaisselle jetable", "Emballages", "Couverts"),
    ),
    "Ingrédients secs": CategorySpec(
        share=0.22, cost_min=3, cost_max=25, margin_min=0.30, margin_max=0.45,
        qty_min=3, qty_max=40, is_equipment=False,
        sub_categories=("Épicerie", "Farines & sucres", "Conserves"),
    ),
    "Hygiène & Entretien": CategorySpec(
        share=0.18, cost_min=4, cost_max=40, margin_min=0.35, margin_max=0.50,
        qty_min=2, qty_max=30, is_equipment=False,
        sub_categories=("Détergents", "Désinfection", "Essuyage"),
    ),
    "Petit équipement": CategorySpec(
        share=0.15, cost_min=30, cost_max=300, margin_min=0.20, margin_max=0.30,
        qty_min=1, qty_max=6, is_equipment=True,
        sub_categories=("Ustensiles pro", "Petit électroménager"),
    ),
    "Gros équipement": CategorySpec(
        share=0.08, cost_min=400, cost_max=3500, margin_min=0.08, margin_max=0.16,
        qty_min=1, qty_max=2, is_equipment=True,
        sub_categories=("Froid professionnel", "Cuisson", "Machines à café"),
    ),
    "Mobilier": CategorySpec(
        share=0.07, cost_min=50, cost_max=800, margin_min=0.12, margin_max=0.22,
        qty_min=1, qty_max=8, is_equipment=True,
        sub_categories=("Tables & chaises", "Rangement", "Comptoirs"),
    ),
}

_SUPPLIERS: tuple[str, ...] = (
    "Promocash", "Metro Pro", "Davigel", "Transgourmet", "Sysco France", "EpiSaveurs",
)

# --- Segmentation client (cf. spec §3.4) ---
# Part de chaque segment + intensité de commande relative (les Grands Comptes
# commandent plus souvent et plus gros).
@dataclass(frozen=True)
class SegmentSpec:
    """Profil d'un segment client.

    Attributes:
        share: Part du segment dans la base clients.
        order_intensity: Poids relatif de fréquence de commande.
        basket_factor: Multiplicateur de taille de panier.
    """

    share: float
    order_intensity: float
    basket_factor: float


SEGMENTS: dict[str, SegmentSpec] = {
    "Grands Comptes": SegmentSpec(share=0.10, order_intensity=3.0, basket_factor=1.8),
    "PME": SegmentSpec(share=0.50, order_intensity=1.5, basket_factor=1.0),
    "Indépendants": SegmentSpec(share=0.40, order_intensity=1.0, basket_factor=0.7),
}

ESTABLISHMENT_TYPES: tuple[str, ...] = ("Restaurant", "Hôtel", "Bar", "Collectivité")

# Cohorte de churn : ~35 clients PME cessent de commander après cette date (cf. spec §3.4).
CHURN_COHORT_SIZE: int = 35
CHURN_CUTOFF: str = "2025-06"


# ===========================================================================
# STRUCTURE PORTEUSE DES DONNÉES PROPRES (avant tout salissage)
# ===========================================================================

@dataclass(frozen=True)
class CleanData:
    """Conteneur immuable des entités propres générées en mémoire.

    Toutes les valeurs y sont parfaitement propres : c'est sur cette base
    qu'est calculée la vérité financière du manifeste (oracle indépendant),
    AVANT que les anomalies ne soient injectées dans les fichiers sources.

    Attributes:
        regions: Référentiel des régions.
        salespeople: Référentiel des commerciaux.
        products: Catalogue produit (avec coût et prix catalogue).
        customers: Référentiel clients.
        orders: En-têtes de commande (rempli au bloc "faits").
        order_lines: Lignes de commande, grain analytique (rempli au bloc "faits").
    """

    regions: pd.DataFrame
    salespeople: pd.DataFrame
    products: pd.DataFrame
    customers: pd.DataFrame
    orders: pd.DataFrame
    order_lines: pd.DataFrame


# ===========================================================================
# GÉNÉRATION DES DIMENSIONS PROPRES
# ===========================================================================

def build_regions() -> pd.DataFrame:
    """Construit le référentiel des régions.

    Returns:
        DataFrame à 5 lignes : ``code_region``, ``nom_region``.
    """
    return pd.DataFrame(
        {"code_region": list(REGIONS.keys()), "nom_region": list(REGIONS.values())}
    )


def build_salespeople(rng: np.random.Generator) -> pd.DataFrame:
    """Construit le référentiel des commerciaux (2 par région).

    La recrue "en montée" reçoit une date d'embauche récente figée
    (cf. ``RISING_SALESPERSON_HIRE``) ; les autres sont réparties entre 2019 et
    2024. Le comportement de montée/déclin lui-même est appliqué plus tard, lors
    de la génération des commandes.

    Args:
        rng: Générateur aléatoire NumPy (déterministe via la graine).

    Returns:
        DataFrame : ``code_commercial``, ``nom``, ``code_region``, ``date_embauche``.
    """
    region_codes = list(REGIONS.keys())
    rows = []
    for i in range(N_SALESPEOPLE):
        code = f"C{i + 1:02d}"
        name = f"{_FIRST_NAMES[i]} {_LAST_NAMES[i]}"
        region = region_codes[i % N_REGIONS]  # 2 commerciaux par région
        if code == RISING_SALESPERSON_CODE:
            hire = pd.Timestamp(RISING_SALESPERSON_HIRE)
        else:
            # Embauche répartie sur ~5 ans, en jours, pour de l'ancienneté variée.
            offset_days = int(rng.integers(0, 365 * 5))
            hire = pd.Timestamp("2019-01-01") + pd.Timedelta(days=offset_days)
        rows.append((code, name, region, hire))

    return pd.DataFrame(
        rows, columns=["code_commercial", "nom", "code_region", "date_embauche"]
    )


def build_products(rng: np.random.Generator) -> pd.DataFrame:
    """Construit le catalogue produit avec des profils de marge contrastés.

    Le prix catalogue est dérivé du coût et d'un taux de marge cible :
    ``prix = cout / (1 - taux_marge)``. La marge n'est donc jamais stockée
    explicitement : elle se recalcule en SQL côté analyse (bonne pratique
    d'audit et mise en valeur du SQL).

    Args:
        rng: Générateur aléatoire NumPy.

    Returns:
        DataFrame : ``code_produit``, ``libelle``, ``categorie``,
        ``sous_categorie``, ``cout_unitaire``, ``prix_catalogue``, ``fournisseur``.
    """
    # Répartition des 120 produits selon la part de chaque catégorie.
    cat_names = list(CATEGORIES.keys())
    shares = np.array([CATEGORIES[c].share for c in cat_names])
    counts = np.round(shares / shares.sum() * N_PRODUCTS).astype(int)
    # Ajustement pour retomber exactement sur N_PRODUCTS malgré les arrondis.
    counts[0] += N_PRODUCTS - counts.sum()

    rows = []
    pid = 1
    for cat_name, n in zip(cat_names, counts):
        spec = CATEGORIES[cat_name]
        for _ in range(int(n)):
            cost = round(float(rng.uniform(spec.cost_min, spec.cost_max)), 2)
            margin = float(rng.uniform(spec.margin_min, spec.margin_max))
            price = round(cost / (1.0 - margin), 2)  # marge cible => prix catalogue
            sub = spec.sub_categories[int(rng.integers(len(spec.sub_categories)))]
            supplier = _SUPPLIERS[int(rng.integers(len(_SUPPLIERS)))]
            rows.append(
                (f"P{pid:04d}", f"{sub} réf.{pid:04d}", cat_name, sub,
                 cost, price, supplier)
            )
            pid += 1

    return pd.DataFrame(
        rows,
        columns=["code_produit", "libelle", "categorie", "sous_categorie",
                 "cout_unitaire", "prix_catalogue", "fournisseur"],
    )


def build_customers(rng: np.random.Generator, salespeople: pd.DataFrame) -> pd.DataFrame:
    """Construit le référentiel clients (segments, régions, commerciaux).

    Chaque client est rattaché à une région, à un commercial de cette région,
    et à un segment tiré selon les parts définies. La date de première commande
    est répartie sur 2023-2025 pour permettre des analyses de cohortes.

    Args:
        rng: Générateur aléatoire NumPy.
        salespeople: Référentiel des commerciaux déjà généré (injecté pour
            éviter de reconsommer le flux aléatoire et garantir la cohérence).

    Returns:
        DataFrame : ``code_client``, ``raison_sociale``, ``segment``,
        ``type_etablissement``, ``code_region``, ``ville``, ``code_commercial``,
        ``date_premiere_commande``.
    """
    region_codes = list(REGIONS.keys())
    seg_names = list(SEGMENTS.keys())
    seg_probs = np.array([SEGMENTS[s].share for s in seg_names])
    seg_probs = seg_probs / seg_probs.sum()

    # Commerciaux disponibles par région (2 chacun), à partir de la table injectée.
    sp_by_region = {
        r: salespeople.loc[salespeople["code_region"] == r, "code_commercial"].tolist()
        for r in region_codes
    }

    rows = []
    for i in range(N_CUSTOMERS):
        code = f"CL{i + 1:04d}"
        region = region_codes[int(rng.integers(N_REGIONS))]
        segment = seg_names[int(rng.choice(len(seg_names), p=seg_probs))]
        etab = ESTABLISHMENT_TYPES[int(rng.integers(len(ESTABLISHMENT_TYPES)))]
        city = CITIES_BY_REGION[region][int(rng.integers(len(CITIES_BY_REGION[region])))]
        salesperson = sp_by_region[region][int(rng.integers(len(sp_by_region[region])))]
        # Date de première commande répartie sur ~3 ans (2023 -> 2025).
        offset_days = int(rng.integers(0, 365 * 3))
        first_order = pd.Timestamp("2023-01-01") + pd.Timedelta(days=offset_days)
        raison = f"{etab} {city} {code[-3:]}"
        rows.append(
            (code, raison, segment, etab, region, city, salesperson, first_order)
        )

    return pd.DataFrame(
        rows,
        columns=["code_client", "raison_sociale", "segment", "type_etablissement",
                 "code_region", "ville", "code_commercial", "date_premiere_commande"],
    )


# ===========================================================================
# GÉNÉRATION DES FAITS (commandes + lignes) AVEC SIGNAUX MÉTIER
# ===========================================================================

# Multiplicateurs de saisonnalité mensuels (index 0 = janvier ... 11 = décembre).
# Logique CHR : creux hivernal, pics en été (terrasses, tourisme) et en décembre.
MONTHLY_SEASONALITY: tuple[float, ...] = (
    0.80, 0.85, 0.95, 1.00, 1.05, 1.20, 1.30, 1.25, 1.05, 1.00, 0.95, 1.25,
)

# Nombre de commandes par mois "de base" (avant saisonnalité et croissance YoY).
# Calibré pour atteindre ~17 000 commandes sur 24 mois.
BASE_ORDERS_PER_MONTH: int = 650

# Statut des commandes : la majorité est livrée ; les annulées sont exclues du CA,
# les retours sont conservés en drapeau (analyse "brut vs net" laissée au Projet 2).
ORDER_STATUS: tuple[str, ...] = ("Livrée", "Retour", "Annulée")
ORDER_STATUS_PROBS: tuple[float, ...] = (0.92, 0.05, 0.03)

# Glissement de mix : part des lignes "équipement" dans le temps.
# Quasi stable jusqu'à mi-2025 (~0,22 -> 0,25), puis rampe marquée au S2 2025
# (~0,25 -> 0,35). C'est ce qui fait monter le CA tout en érodant la marge.
EQUIP_SHARE_START: float = 0.22
EQUIP_SHARE_MID: float = 0.25
EQUIP_SHARE_END: float = 0.35
MID_MONTH_INDEX: int = 18  # 2025-07 : début de la rampe du S2 2025

# Fuite de marge : 3 produits à marge élevée sur-remisés à la commande.
MARGIN_LEAK_N: int = 3
MARGIN_LEAK_REMISE_RANGE: tuple[float, float] = (0.20, 0.40)
NORMAL_REMISE_RANGE: tuple[float, float] = (0.0, 0.08)
HIGH_MARGIN_CATEGORIES: tuple[str, ...] = (
    "Consommables", "Ingrédients secs", "Hygiène & Entretien",
)

# Nombre de lignes par commande selon le segment (les Grands Comptes ont de plus
# gros paniers). Bornes au sens de ``rng.integers`` (borne haute exclue).
LINES_PER_ORDER_BY_SEGMENT: dict[str, tuple[int, int]] = {
    "Grands Comptes": (3, 9),
    "PME": (1, 7),
    "Indépendants": (1, 5),
}


def select_margin_leak_products(
    products: pd.DataFrame, rng: np.random.Generator
) -> list[str]:
    """Désigne les produits victimes de la fuite de marge (sur-remise).

    On choisit ``MARGIN_LEAK_N`` produits parmi les catégories à marge élevée :
    le sur-discount y est économiquement le plus parlant (on rogne une marge
    qui aurait dû être confortable).

    Args:
        products: Catalogue produit.
        rng: Générateur aléatoire NumPy.

    Returns:
        Liste de codes produits sur-remisés.
    """
    pool = products.loc[
        products["categorie"].isin(HIGH_MARGIN_CATEGORIES), "code_produit"
    ].to_numpy()
    chosen = rng.choice(pool, size=MARGIN_LEAK_N, replace=False)
    return [str(c) for c in chosen]


def select_churn_cohort(customers: pd.DataFrame) -> list[str]:
    """Désigne la cohorte de clients PME qui vont cesser de commander.

    On retient les PME ayant la plus ancienne date de première commande :
    ce sont des clients installés dont le départ est un signal de rétention
    crédible (et non un simple bruit).

    Args:
        customers: Référentiel clients.

    Returns:
        Liste de codes clients (cohorte de churn).
    """
    pme = customers[customers["segment"] == "PME"].sort_values(
        "date_premiere_commande"
    )
    return pme["code_client"].head(CHURN_COHORT_SIZE).tolist()


def equipment_share(month_index: int) -> float:
    """Retourne la part cible de lignes "équipement" pour un mois donné.

    Rampe en deux temps : douce jusqu'à mi-2025, puis marquée au S2 2025.

    Args:
        month_index: Index du mois (0 = premier mois de la période).

    Returns:
        Part attendue de lignes équipement (entre 0 et 1).
    """
    if month_index <= MID_MONTH_INDEX:
        # Progression douce du début jusqu'à 2025-06.
        frac = month_index / MID_MONTH_INDEX
        return EQUIP_SHARE_START + (EQUIP_SHARE_MID - EQUIP_SHARE_START) * frac
    # Rampe marquée sur le second semestre 2025.
    frac = (month_index - MID_MONTH_INDEX) / (23 - MID_MONTH_INDEX)
    return EQUIP_SHARE_MID + (EQUIP_SHARE_END - EQUIP_SHARE_MID) * frac


def _salesperson_time_factor(code: str, month_index: int) -> float:
    """Facteur d'activité d'un commercial selon le mois (signaux montée/déclin).

    - Recrue en montée : ~0 avant l'embauche, puis rampe vers 1 sur ~6 mois.
    - Commercial en déclin : décroît linéairement de 1,0 à 0,7 sur 24 mois.
    - Autres : 1,0 (constant).

    Args:
        code: Code commercial.
        month_index: Index du mois.

    Returns:
        Multiplicateur de poids appliqué aux clients de ce commercial.
    """
    if code == RISING_SALESPERSON_CODE:
        hire_index = 8  # 2024-09 = 9e mois => index 8
        return float(np.clip((month_index - hire_index) / 6.0, 0.0, 1.0))
    if code == DECLINING_SALESPERSON_CODE:
        return 1.0 - 0.30 * (month_index / 23.0)
    return 1.0


def build_orders_and_lines(
    rng: np.random.Generator,
    customers: pd.DataFrame,
    products: pd.DataFrame,
    leak_products: list[str],
    churn_customers: list[str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Génère les commandes et leurs lignes en y imprimant tous les signaux.

    Signaux injectés : saisonnalité, croissance YoY, glissement de mix vers
    l'équipement, fuite de marge, churn de cohorte, trajectoires
    montée/déclin des commerciaux.

    Complexité : O(nb_commandes × lignes_par_commande). À volumétrie 100x, la
    boucle Python sur les lignes deviendrait le goulot : on la remplacerait par
    un tirage vectorisé NumPy (génération des lignes en bloc puis affectation
    aux commandes), sans changer la logique.

    Args:
        rng: Générateur aléatoire NumPy.
        customers: Référentiel clients.
        products: Catalogue produit.
        leak_products: Codes produits sur-remisés (fuite de marge).
        churn_customers: Codes clients de la cohorte de churn.

    Returns:
        Couple (orders, order_lines).
    """
    months = pd.period_range(PERIOD_START, PERIOD_END, freq="M")
    leak_set = set(leak_products)
    churn_set = set(churn_customers)
    churn_cutoff = pd.Period(CHURN_CUTOFF, freq="M")

    # --- Pré-calculs clients (vectorisés) pour pondérer les tirages ---
    cust_codes = customers["code_client"].to_numpy()
    cust_segment = customers["segment"].to_numpy()
    cust_rep = customers["code_commercial"].to_numpy()
    cust_basket = np.array(
        [SEGMENTS[s].basket_factor for s in cust_segment], dtype=float
    )
    base_intensity = np.array(
        [SEGMENTS[s].order_intensity for s in cust_segment], dtype=float
    )
    is_churn = np.array([c in churn_set for c in cust_codes], dtype=bool)

    # --- Pré-calculs produits par famille (équipement vs reste) ---
    equip_cats = [c for c, s in CATEGORIES.items() if s.is_equipment]
    non_equip_cats = [c for c, s in CATEGORIES.items() if not s.is_equipment]

    def _pool(categories: list[str]) -> dict[str, np.ndarray]:
        sub = products[products["categorie"].isin(categories)]
        return {
            "code": sub["code_produit"].to_numpy(),
            "price": sub["prix_catalogue"].to_numpy(),
            "cat": sub["categorie"].to_numpy(),
        }

    equip_pool = _pool(equip_cats)
    non_equip_pool = _pool(non_equip_cats)
    # Bornes de quantité par produit (indexées par code) pour tirer la quantité.
    qty_bounds = {
        row.code_produit: (CATEGORIES[row.categorie].qty_min,
                           CATEGORIES[row.categorie].qty_max)
        for row in products.itertuples()
    }

    order_rows: list[tuple] = []
    line_rows: list[tuple] = []
    order_counter = 0
    line_counter = 0

    for m_idx, period in enumerate(months):
        # 1) Combien de commandes ce mois-ci ? (saisonnalité × croissance YoY)
        season = MONTHLY_SEASONALITY[period.month - 1]
        yoy = YOY_GROWTH if period.year == 2025 else 1.0
        n_orders = int(round(BASE_ORDERS_PER_MONTH * season * yoy))

        # 2) Poids de chaque client ce mois (intensité × commercial × churn)
        rep_factor = np.array(
            [_salesperson_time_factor(r, m_idx) for r in cust_rep], dtype=float
        )
        churn_factor = np.where(is_churn & (period > churn_cutoff), 0.0, 1.0)
        weights = base_intensity * rep_factor * churn_factor
        total = weights.sum()
        if total <= 0:
            continue
        probs = weights / total

        # 3) Tirage des clients de ce mois (avec remise)
        idx = rng.choice(len(cust_codes), size=n_orders, p=probs)
        days_in_month = period.days_in_month
        equip_target = equipment_share(m_idx)

        for j in idx:
            order_counter += 1
            order_id = f"CMD{period.year}{period.month:02d}{order_counter:06d}"
            day = int(rng.integers(0, days_in_month))
            order_date = period.start_time + pd.Timedelta(days=day)
            status = ORDER_STATUS[
                int(rng.choice(len(ORDER_STATUS), p=ORDER_STATUS_PROBS))
            ]
            order_rows.append(
                (order_id, order_date, cust_codes[j], cust_rep[j], status)
            )

            # 4) Lignes de la commande
            lo, hi = LINES_PER_ORDER_BY_SEGMENT[cust_segment[j]]
            n_lines = int(rng.integers(lo, hi))
            basket = cust_basket[j]
            for _ in range(n_lines):
                # Équipement ou non, selon la part cible du mois.
                if rng.random() < equip_target:
                    pool = equip_pool
                else:
                    pool = non_equip_pool
                k = int(rng.integers(len(pool["code"])))
                code_prod = str(pool["code"][k])
                price = float(pool["price"][k])
                qmin, qmax = qty_bounds[code_prod]
                qty = max(1, int(round(rng.integers(qmin, qmax + 1) * basket)))
                # Remise : normale, sauf produits sur-remisés (fuite de marge).
                if code_prod in leak_set:
                    remise = float(rng.uniform(*MARGIN_LEAK_REMISE_RANGE))
                else:
                    remise = float(rng.uniform(*NORMAL_REMISE_RANGE))
                line_counter += 1
                line_rows.append(
                    (f"L{line_counter:07d}", order_id, code_prod, qty,
                     round(price, 2), round(remise, 3))
                )

    orders = pd.DataFrame(
        order_rows,
        columns=["commande_id", "date_commande", "code_client",
                 "code_commercial", "statut"],
    )
    order_lines = pd.DataFrame(
        line_rows,
        columns=["ligne_id", "commande_id", "code_produit", "quantite",
                 "prix_unitaire_vente", "remise_pct"],
    )
    return orders, order_lines


# ===========================================================================
# SALISSAGE PAR SOURCE (injection des anomalies qualité)
# ===========================================================================
#
# Chaque injecteur prend les données PROPRES et retourne (artefact_sale, compteurs).
# Les compteurs alimenteront le manifeste. Rappel : on injecte ici des DÉFAUTS
# qualité (à nettoyer), distincts des signaux métier (à découvrir) du bloc 2.

# Taux d'anomalies (cf. spec §5). La saleté est surtout systématique : seules les
# lignes manuelles et les doublons d'export sont sporadiques.
SALES_RATE_MANUAL: float = 0.004   # lignes d'ajustement saisies à la main
SALES_RATE_DUP: float = 0.01       # doublons par chevauchement d'export
CRM_RATE_BAD_DATE: float = 0.015   # dates de 1re commande saisies à la main
CRM_RATE_SEG_CASE: float = 0.12    # incohérences de casse sur le segment
CATALOG_VARIANT_SHARE: float = 0.35  # part des produits d'une catégorie ciblée

# En-têtes "sales" de l'export ERP (codes, espaces parasites sur 2 colonnes).
H_QTE: str = "Quantité "      # espace final (gabarit de rapport ERP)
H_PU: str = "PU Vente HT "    # espace final

# Variantes orthographiques de catégorie injectées dans le catalogue.
CATEGORY_VARIANTS: dict[str, tuple[str, ...]] = {
    "Consommables": ("consommables", "Conso."),
    "Hygiène & Entretien": ("hygiene et entretien",),
}


def _format_fr_amount(value: float) -> str:
    """Formate un montant à la française : espace insécable milliers, virgule décimale.

    Reproduit le comportement d'un export ERP sous locale FR (ex: 1234.5 -> '1 234,50').

    Args:
        value: Montant numérique.

    Returns:
        Chaîne au format monétaire français.
    """
    s = f"{value:,.2f}"                       # '1,234.50' (format US)
    return s.replace(",", "\u00a0").replace(".", ",")  # '1 234,50' (FR, nbsp)


def inject_sales_anomalies(
    rng: np.random.Generator, orders: pd.DataFrame, order_lines: pd.DataFrame
) -> tuple[pd.DataFrame, dict]:
    """Salit l'export des ventes (dénormalisé, grain ligne).

    Anomalies : montants au format FR (systématique), en-têtes à espaces
    (systématique), lignes d'ajustement manuelles (montant et parfois date
    invalides, sporadique), doublons d'export (sporadique).

    Args:
        rng: Générateur aléatoire NumPy.
        orders: En-têtes de commande propres.
        order_lines: Lignes de commande propres.

    Returns:
        Couple (df_sale_avec_colonne__month, compteurs). La colonne ``_month``
        sert au découpage en 24 fichiers mensuels par l'écrivain ; elle n'est
        pas écrite dans les fichiers.
    """
    flat = order_lines.merge(orders, on="commande_id")
    flat["_month"] = flat["date_commande"].dt.to_period("M").astype(str)
    n = len(flat)

    # --- Lignes manuelles : montant non numérique, et pour la moitié date impossible ---
    n_manual = int(round(n * SALES_RATE_MANUAL))
    manual_idx = rng.choice(n, size=n_manual, replace=False)
    n_bad_date = n_manual // 2
    bad_date_idx = manual_idx[:n_bad_date]

    # PU au format FR pour les lignes numériques ; "N/A"/"" pour les lignes manuelles.
    pu_str = np.array([_format_fr_amount(v) for v in flat["prix_unitaire_vente"].to_numpy()],
                      dtype=object)
    pu_str[manual_idx] = rng.choice(["N/A", ""], size=n_manual)

    # Dates : JJ/MM/AAAA, sauf dates impossibles sur les lignes manuelles concernées.
    date_str = flat["date_commande"].dt.strftime("%d/%m/%Y").to_numpy().astype(object)
    years = flat["date_commande"].dt.year.to_numpy()
    for i in bad_date_idx:
        date_str[i] = f"32/13/{years[i]}"

    base = pd.DataFrame({
        "date": date_str,
        "commande_id": flat["commande_id"].to_numpy(),
        "ligne": flat["ligne_id"].to_numpy(),
        "code_client": flat["code_client"].to_numpy(),
        "code_produit": flat["code_produit"].to_numpy(),
        "quantite": flat["quantite"].to_numpy(),
        "pu": pu_str,
        "remise": np.round(flat["remise_pct"].to_numpy() * 100, 1),
        "code_commercial": flat["code_commercial"].to_numpy(),
        "statut": flat["statut"].to_numpy(),
        "_month": flat["_month"].to_numpy(),
    })

    # --- Doublons d'export : concentrés sur les 3 premiers/derniers jours du mois ---
    day = flat["date_commande"].dt.day.to_numpy()
    dim = flat["date_commande"].dt.days_in_month.to_numpy()
    boundary_idx = np.where((day <= 3) | (day >= dim - 2))[0]
    n_dup = min(int(round(n * SALES_RATE_DUP)), len(boundary_idx))
    dup_idx = rng.choice(boundary_idx, size=n_dup, replace=False)
    duplicates = base.iloc[dup_idx].copy()

    dirty = pd.concat([base, duplicates], ignore_index=True)
    dirty = dirty.rename(columns={
        "date": "Date Commande", "commande_id": "N° Commande",
        "ligne": "N° Ligne",
        "code_client": "Code Client", "code_produit": "Code Produit",
        "quantite": H_QTE, "pu": H_PU, "remise": "Remise %",
        "code_commercial": "Code Commercial", "statut": "Statut",
    })

    # CA des lignes manuelles LIVRÉES : irréversiblement perdu (montant -> NaN).
    # Sert à calculer le CA réconcilié attendu (truth - perte) dans le manifeste.
    clean_ca = (flat["quantite"] * flat["prix_unitaire_vente"]
                * (1 - flat["remise_pct"])).to_numpy()
    is_livree = (flat["statut"] == "Livrée").to_numpy()
    ca_perdue = float(clean_ca[manual_idx][is_livree[manual_idx]].sum())

    counters = {
        "montants_format_fr": {"nature": "systematique", "nb_lignes": int(n - n_manual)},
        "entetes_espaces": {"colonnes": [H_QTE, H_PU]},
        "lignes_manuelles_montant_invalide": {
            "nb": int(n_manual),
            "ca_perdue_livree_eur": round(ca_perdue, 2),
        },
        "lignes_manuelles_date_invalide": {"nb": int(n_bad_date)},
        "doublons_export": {
            "nb_lignes": int(n_dup),
            "commandes_concernees": base.iloc[dup_idx]["commande_id"].head(10).tolist(),
        },
    }
    return dirty, counters


def inject_crm_anomalies(
    rng: np.random.Generator, customers: pd.DataFrame
) -> tuple[pd.DataFrame, dict]:
    """Salit le référentiel client (saisie commerciale).

    Anomalies : encodage cp1252 (géré à l'écriture), dates de première commande
    invalides (sporadique), incohérences de casse sur le segment (modéré).

    Args:
        rng: Générateur aléatoire NumPy.
        customers: Référentiel clients propre.

    Returns:
        Couple (df_sale, compteurs).
    """
    df = customers.copy()
    n = len(df)

    # Dates en JJ/MM/AAAA, avec une fraction de dates impossibles (saisie humaine).
    date_str = df["date_premiere_commande"].dt.strftime("%d/%m/%Y").to_numpy().astype(object)
    years = df["date_premiere_commande"].dt.year.to_numpy()
    n_bad = int(round(n * CRM_RATE_BAD_DATE))
    bad_idx = rng.choice(n, size=n_bad, replace=False)
    for i in bad_idx:
        date_str[i] = f"32/13/{years[i]}"
    df["date_premiere_commande"] = date_str

    # Casse/espaces incohérents sur le segment (champ libre).
    seg = df["segment"].to_numpy().astype(object)
    n_seg = int(round(n * CRM_RATE_SEG_CASE))
    seg_idx = rng.choice(n, size=n_seg, replace=False)
    for i in seg_idx:
        original = seg[i]
        seg[i] = str(rng.choice([original.lower(), original.upper(), original + " "]))
    df["segment"] = seg

    counters = {
        "encodage": "cp1252",
        "date_premiere_commande_invalide": {"nb": int(n_bad)},
        "segment_casse_variante": {"nb_lignes": int(n_seg)},
    }
    return df, counters


def inject_catalog_anomalies(
    rng: np.random.Generator, products: pd.DataFrame
) -> tuple[pd.DataFrame, dict]:
    """Salit le catalogue produit (référentiel maintenu à la main).

    Anomalie : doublons orthographiques de catégorie sur 2 catégories.
    Non corrigée, elle éclate une catégorie en plusieurs et fausse l'effet mix.

    Args:
        rng: Générateur aléatoire NumPy.
        products: Catalogue propre.

    Returns:
        Couple (df_sale, compteurs).
    """
    df = products.copy()
    cat = df["categorie"].to_numpy().astype(object)
    impacted = 0
    for target, variants in CATEGORY_VARIANTS.items():
        idx = np.where(df["categorie"].to_numpy() == target)[0]
        n_var = int(round(len(idx) * CATALOG_VARIANT_SHARE))
        chosen = rng.choice(idx, size=n_var, replace=False)
        for i in chosen:
            cat[i] = str(rng.choice(list(variants)))
            impacted += 1
    df["categorie"] = cat

    counters = {
        "categories_variantes": {
            **{k: list(v) for k, v in CATEGORY_VARIANTS.items()},
            "nb_produits_impactes": int(impacted),
        }
    }
    return df, counters


def build_budget_long(
    rng: np.random.Generator,
    orders: pd.DataFrame,
    order_lines: pd.DataFrame,
    products: pd.DataFrame,
    customers: pd.DataFrame,
    regions: pd.DataFrame,
) -> tuple[pd.DataFrame, dict]:
    """Construit le budget (CA et marge budgétés par mois × région × catégorie).

    Le budget est calé proche du réel (facteur ~U(0,95 ; 1,10)) SAUF pour la
    région Sud, volontairement sur-budgétée (×U(1,15 ; 1,25)) : c'est le signal
    "Sud sous budget" (le réel y reste sous l'objectif). Ce n'est pas une
    anomalie qualité mais une réalité de gestion à analyser.

    La forme "large" pénible (mois en colonnes, titre, sous-totaux) est appliquée
    à l'écriture, pas ici : on retourne un format long propre + le réel pour audit.

    Args:
        rng: Générateur aléatoire NumPy.
        orders, order_lines, products, customers: Données propres.
        regions: Référentiel des régions (pour le libellé).

    Returns:
        Couple (df_long, compteurs). ``df_long`` contient le budget ET le réel
        (colonnes ``ca_reel``/``marge_reel`` utiles à l'audit, non écrites).
    """
    enr = (
        order_lines.merge(
            orders[["commande_id", "date_commande", "code_client", "statut"]],
            on="commande_id")
        .query("statut == 'Livrée'")
        .merge(products[["code_produit", "categorie", "cout_unitaire"]], on="code_produit")
        .merge(customers[["code_client", "code_region"]], on="code_client")
    )
    enr["ca"] = enr["quantite"] * enr["prix_unitaire_vente"] * (1 - enr["remise_pct"])
    enr["marge"] = enr["quantite"] * (
        enr["prix_unitaire_vente"] * (1 - enr["remise_pct"]) - enr["cout_unitaire"]
    )
    enr["periode"] = enr["date_commande"].dt.to_period("M").astype(str)

    real = (
        enr.groupby(["periode", "code_region", "categorie"])
        .agg(ca_reel=("ca", "sum"), marge_reel=("marge", "sum"))
        .reset_index()
    )

    factor = rng.uniform(0.93, 1.07, size=len(real))
    sud_mask = real["code_region"].to_numpy() == UNDERBUDGET_REGION_CODE
    factor[sud_mask] = rng.uniform(1.15, 1.25, size=int(sud_mask.sum()))
    real["ca_budgete"] = (real["ca_reel"] * factor).round(0)
    real["marge_budgetee"] = (real["marge_reel"] * factor).round(0)

    long_df = real.merge(regions, on="code_region")
    counters = {
        "format": "large",
        "sous_totaux_intercales": True,
        "signal_sud_sous_budget": True,
    }
    return long_df, counters

# ===========================================================================
# ORACLE, ÉCRITURE DES SOURCES, MANIFESTE, ORCHESTRATION
# ===========================================================================

import json
import logging
from datetime import datetime, timezone
from pathlib import Path

# Répertoires de sortie par défaut (surchargeables via main()).
RAW_DIR_DEFAULT: Path = Path("data_raw")
MANIFEST_NAME: str = "_manifest_anomalies.json"

# Libellés de mois FR pour la mise en page large du budget.
MONTH_LABELS_FR: tuple[str, ...] = (
    "Janv", "Févr", "Mars", "Avr", "Mai", "Juin",
    "Juil", "Août", "Sept", "Oct", "Nov", "Déc",
)


def compute_truth(
    orders: pd.DataFrame, order_lines: pd.DataFrame, products: pd.DataFrame
) -> dict:
    """Calcule la vérité financière sur les données PROPRES (oracle indépendant).

    Point critique : cette fonction ne voit jamais les fichiers sales. Elle est
    le chemin de calcul de référence, distinct du pipeline de nettoyage, ce qui
    rend la réconciliation probante (et non circulaire).

    Args:
        orders: En-têtes de commande propres.
        order_lines: Lignes de commande propres.
        products: Catalogue propre (pour le coût).

    Returns:
        Dictionnaire de vérité (CA livré, marge, taux, comptes, nb catégories).
    """
    lines = order_lines.merge(
        products[["code_produit", "cout_unitaire"]], on="code_produit"
    ).merge(orders[["commande_id", "statut"]], on="commande_id")
    lines["ca"] = lines["quantite"] * lines["prix_unitaire_vente"] * (1 - lines["remise_pct"])
    lines["marge"] = lines["quantite"] * (
        lines["prix_unitaire_vente"] * (1 - lines["remise_pct"]) - lines["cout_unitaire"]
    )
    liv = lines[lines["statut"] == "Livrée"]
    ca = float(liv["ca"].sum())
    marge = float(liv["marge"].sum())
    return {
        "ca_total_livre_eur": round(ca, 2),
        "marge_total_livree_eur": round(marge, 2),
        "taux_marge_global_pct": round(marge / ca * 100, 2),
        "nb_commandes": int(len(orders)),
        "nb_lignes_total": int(len(order_lines)),
        "nb_lignes_livrees": int(len(liv)),
        "nb_categories_canoniques": int(len(CATEGORIES)),
    }


# --- Écrivains de sources -------------------------------------------------

def write_sales(sales_dirty: pd.DataFrame, raw_dir: Path) -> int:
    """Écrit l'export des ventes en 24 fichiers mensuels (.xlsx).

    Args:
        sales_dirty: Ventes salies, avec colonne ``_month`` de découpage.
        raw_dir: Répertoire de sortie.

    Returns:
        Nombre de fichiers écrits.
    """
    cols = [c for c in sales_dirty.columns if c != "_month"]
    count = 0
    for month, grp in sales_dirty.groupby("_month"):
        grp[cols].to_excel(raw_dir / f"export_ventes_{month}.xlsx", index=False)
        count += 1
    return count


def write_crm(crm_dirty: pd.DataFrame, raw_dir: Path) -> None:
    """Écrit le référentiel client en CSV ``cp1252`` (réalisme vieux CRM Windows)."""
    crm_dirty.to_csv(
        raw_dir / "export_clients_crm.csv", sep=";", index=False, encoding="cp1252"
    )


def write_catalog(catalog_dirty: pd.DataFrame, raw_dir: Path) -> None:
    """Écrit le catalogue produit (.xlsx)."""
    catalog_dirty.to_excel(raw_dir / "catalogue_produits.xlsx", index=False)


def write_salespeople(salespeople: pd.DataFrame, raw_dir: Path) -> None:
    """Écrit le référentiel commerciaux (.csv utf-8, source propre de contraste)."""
    df = salespeople.copy()
    df["date_embauche"] = df["date_embauche"].dt.strftime("%Y-%m-%d")
    df.to_csv(raw_dir / "commerciaux.csv", sep=";", index=False, encoding="utf-8")


def _write_budget_sheet(ws: Any, value_map: dict, title: str) -> None:
    """Écrit une feuille de budget au format "large" pénible (titre, fusions, sous-totaux).

    Reproduit un fichier de contrôle de gestion : ligne de titre fusionnée,
    région fusionnée verticalement sur ses catégories, mois en colonnes,
    sous-totaux par région intercalés. Le nettoyage côté Projet 1 = un reshape.
    """
    from openpyxl.styles import Font, Alignment

    n_cols = 2 + 12
    ws.cell(1, 1, title).font = Font(name="Arial", bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    headers = ["Région", "Catégorie", *MONTH_LABELS_FR]
    for j, h in enumerate(headers, start=1):
        ws.cell(3, j, h).font = Font(name="Arial", bold=True)

    region_order = list(REGIONS.values())
    cat_order = list(CATEGORIES.keys())
    r = 4
    for region in region_order:
        start = r
        for cat in cat_order:
            ws.cell(r, 1, region if cat == cat_order[0] else None)
            ws.cell(r, 2, cat)
            for m in range(1, 13):
                ws.cell(r, 2 + m, round(value_map.get((region, cat, m), 0.0)))
            r += 1
        ws.merge_cells(start_row=start, start_column=1, end_row=r - 1, end_column=1)
        ws.cell(start, 1).alignment = Alignment(vertical="top")
        # Sous-total région (valeurs en dur : fichier de données brut, pas un modèle).
        ws.cell(r, 2, "Sous-total").font = Font(name="Arial", bold=True)
        for m in range(1, 13):
            total = sum(value_map.get((region, cat, m), 0.0) for cat in cat_order)
            ws.cell(r, 2 + m, round(total)).font = Font(name="Arial", bold=True)
        r += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22


def write_budget_wide(budget_long: pd.DataFrame, raw_dir: Path) -> None:
    """Écrit les 2 fichiers budget annuels au format large (feuilles CA et Marge)."""
    from openpyxl import Workbook

    b = budget_long.copy()
    b["year"] = b["periode"].str[:4].astype(int)
    b["month"] = b["periode"].str[5:7].astype(int)

    sheets = [("CA", "ca_budgete", "CA budgété (€)"),
              ("Marge", "marge_budgetee", "Marge budgétée (€)")]
    for year in (2024, 2025):
        wb = Workbook()
        for i, (sheet_name, metric, label) in enumerate(sheets):
            ws = wb.active if i == 0 else wb.create_sheet(sheet_name)
            ws.title = sheet_name
            sub = b[b["year"] == year]
            vmap = {
                (row.nom_region, row.categorie, row.month): float(getattr(row, metric))
                for row in sub.itertuples()
            }
            _write_budget_sheet(ws, vmap, f"Budget commercial {year} — {label}")
        wb.save(raw_dir / f"budget_{year}.xlsx")


def write_manifest(
    truth: dict, c_sales: dict, c_crm: dict, c_catalog: dict, c_budget: dict,
    manifest_path: Path,
) -> dict:
    """Assemble et écrit le manifeste (vérité + anomalies + attendus après nettoyage).

    Les "attendus" tiennent compte de la perte irréversible : les lignes
    manuelles à montant invalide ne sont pas récupérables, donc le CA réconcilié
    attendu = CA vrai - CA perdu sur ces lignes (livrées).

    Returns:
        Le dictionnaire manifeste écrit.
    """
    ca_perdue = c_sales["lignes_manuelles_montant_invalide"]["ca_perdue_livree_eur"]
    expected = {
        "ca_reconcilie_attendu_eur": round(truth["ca_total_livre_eur"] - ca_perdue, 2),
        "nb_lignes_apres_dedup": truth["nb_lignes_total"],
        "nb_dates_invalides_attendu": c_sales["lignes_manuelles_date_invalide"]["nb"],
        "nb_montants_invalides_attendu": c_sales["lignes_manuelles_montant_invalide"]["nb"],
        "nb_categories_apres_normalisation": truth["nb_categories_canoniques"],
    }
    manifest = {
        "generation": {
            "seed": SEED,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "periode": f"{PERIOD_START}..{PERIOD_END}",
            "version_generateur": "1.1.0",
        },
        "verite_financiere": truth,
        "anomalies_injectees": {
            "ventes": c_sales, "crm": c_crm, "catalogue": c_catalog, "budget": c_budget,
        },
        "attendus_apres_nettoyage": expected,
    }
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return manifest


def setup_logging() -> logging.Logger:
    """Configure un logger console simple pour le suivi d'exécution."""
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s",
    )
    return logging.getLogger("generateur")


def generate_dataset(raw_dir: Path, manifest_path: Path) -> dict:
    """Orchestre toute la génération de façon déterministe (graine fixe).

    Args:
        raw_dir: Répertoire des fichiers sources.
        manifest_path: Chemin du manifeste.

    Returns:
        Le manifeste écrit.
    """
    log = setup_logging()
    raw_dir.mkdir(parents=True, exist_ok=True)
    rng = np.random.default_rng(SEED)

    log.info("Dimensions...")
    regions = build_regions()
    salespeople = build_salespeople(rng)
    products = build_products(rng)
    customers = build_customers(rng, salespeople)
    leak = select_margin_leak_products(products, rng)
    churn = select_churn_cohort(customers)

    log.info("Faits (commandes + lignes + signaux)...")
    orders, order_lines = build_orders_and_lines(rng, customers, products, leak, churn)

    log.info("Vérité (oracle, sur données propres)...")
    truth = compute_truth(orders, order_lines, products)

    log.info("Salissage des sources...")
    sales_dirty, c_sales = inject_sales_anomalies(rng, orders, order_lines)
    crm_dirty, c_crm = inject_crm_anomalies(rng, customers)
    catalog_dirty, c_catalog = inject_catalog_anomalies(rng, products)
    budget_long, c_budget = build_budget_long(
        rng, orders, order_lines, products, customers, regions
    )

    log.info("Écriture des fichiers...")
    n_sales = write_sales(sales_dirty, raw_dir)
    write_crm(crm_dirty, raw_dir)
    write_catalog(catalog_dirty, raw_dir)
    write_salespeople(salespeople, raw_dir)
    write_budget_wide(budget_long, raw_dir)
    manifest = write_manifest(truth, c_sales, c_crm, c_catalog, c_budget, manifest_path)

    log.info("Terminé : %d ventes + CRM + catalogue + 2 budgets + commerciaux + manifeste",
             n_sales)
    log.info("Vérité : CA livré = %.2f M€ | marge = %.1f %% | %d commandes | %d lignes",
             truth["ca_total_livre_eur"] / 1e6, truth["taux_marge_global_pct"],
             truth["nb_commandes"], truth["nb_lignes_total"])
    return manifest


def main() -> None:
    """Point d'entrée : génère le dataset dans ``data_raw/`` + manifeste."""
    generate_dataset(RAW_DIR_DEFAULT, RAW_DIR_DEFAULT / MANIFEST_NAME)


# ===========================================================================
# VÉRIFICATION DE BOUCLAGE (preuve que l'oracle est exploitable)
# ===========================================================================

def verify_roundtrip(raw_dir: Path, manifest: dict) -> None:
    """Relit les fichiers sales, applique un nettoyage minimal, et prouve que
    l'on retombe exactement sur les valeurs attendues du manifeste.

    C'est la démonstration que le manifeste fait office d'oracle exploitable,
    avant même que le pipeline complet du Projet 1 existe.
    """
    exp = manifest["attendus_apres_nettoyage"]

    # 1) Lecture + concat des 24 ventes, normalisation des en-têtes.
    files = sorted(raw_dir.glob("export_ventes_*.xlsx"))
    v = pd.concat([pd.read_excel(f, dtype=str) for f in files], ignore_index=True)
    v.columns = [c.strip() for c in v.columns]

    # 2) Déduplication sur la clé naturelle (N° Commande, N° Ligne).
    n_before = len(v)
    v = v.drop_duplicates(subset=["N° Commande", "N° Ligne"])
    n_after = len(v)

    # 3) Parsing des montants FR (espaces + virgule -> point) et des dates.
    pu = (v["PU Vente HT"].astype("string")
          .str.replace("\u00a0", "", regex=False)
          .str.replace(" ", "", regex=False)
          .str.replace(",", ".", regex=False))
    pu = pd.to_numeric(pu, errors="coerce")
    n_montant_invalid = int(pu.isna().sum())

    d = pd.to_datetime(v["Date Commande"], format="%d/%m/%Y", errors="coerce")
    n_date_invalid = int(d.isna().sum())

    qte = pd.to_numeric(v["Quantité"], errors="coerce")
    rem = pd.to_numeric(v["Remise %"], errors="coerce") / 100.0
    work = pd.DataFrame({"pu": pu, "d": d, "qte": qte, "rem": rem, "statut": v["Statut"]})
    work = work[work["d"].notna()]  # lignes à date invalide écartées
    liv = work[work["statut"] == "Livrée"]
    ca = float((liv["qte"] * liv["pu"] * (1 - liv["rem"])).sum())  # NaN (montant KO) ignorés

    # 4) Normalisation des catégories du catalogue.
    catalog = pd.read_excel(raw_dir / "catalogue_produits.xlsx")
    canon = {var: target for target, vs in CATEGORY_VARIANTS.items() for var in vs}
    n_cat = catalog["categorie"].map(lambda x: canon.get(x, x)).nunique()

    # 5) Confrontation à l'oracle.
    gap = abs(ca - exp["ca_reconcilie_attendu_eur"])
    print("=" * 72)
    print("VÉRIFICATION DE BOUCLAGE (pipeline minimal vs manifeste)")
    print("=" * 72)
    checks = [
        ("Lignes après dédup", n_after, exp["nb_lignes_apres_dedup"]),
        ("Dates invalides", n_date_invalid, exp["nb_dates_invalides_attendu"]),
        ("Montants invalides", n_montant_invalid, exp["nb_montants_invalides_attendu"]),
        ("Catégories canoniques", n_cat, exp["nb_categories_apres_normalisation"]),
    ]
    for label, got, want in checks:
        status = "OK " if got == want else "ÉCHEC"
        print(f"[{status}] {label:<24} obtenu={got:<12} attendu={want}")
    print(f"[{'OK ' if gap < 0.01 else 'ÉCHEC'}] {'CA réconcilié (€)':<24} "
          f"obtenu={ca:,.2f}  attendu={exp['ca_reconcilie_attendu_eur']:,.2f}  écart={gap:.4f}")
    print(f"\nDoublons retirés : {n_before - n_after} | fichiers ventes lus : {len(files)}")

    all_ok = all(g == w for _, g, w in checks) and gap < 0.01
    print("\n>>> ORACLE VALIDÉ : bouclage exact." if all_ok
          else "\n>>> ÉCHEC : divergence détectée.")


if __name__ == "__main__":
    out = Path("data_raw")
    manifest = generate_dataset(out, out / MANIFEST_NAME)
    print()
    verify_roundtrip(out, manifest)